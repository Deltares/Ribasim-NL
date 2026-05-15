# %%
"""Script met verschillende functies om de uitvoer van de Ribasim modellen te vergelijken met meetreeksen"""

import ast
import operator
from collections import defaultdict
from pathlib import Path
from typing import Any

import geopandas as gpd
import matplotlib.patheffects as pe
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import tqdm
import xarray as xr
from openpyxl.styles import PatternFill
from shapely import wkt

from ribasim_nl import CloudStorage
from ribasim_nl.aquo import waterbeheercode
from ribasim_nl.html_viewer import CreateHTMLViewer
from ribasim_nl.model import Model

try:
    from ribasim_nl.assign_lhm_fractions import get_lhm_fractions
except ImportError:
    get_lhm_fractions = None  # type: ignore[assignment]


def _sanitize_filename(name: str) -> str:
    r"""Verwijdert of vervangt tekens die niet geldig zijn in bestandsnamen op Windows/Linux.

    Spaties worden underscores, alle overige ongeldige tekens (/ \\ : * ? " < > | ( ) , )
    worden weggelaten.
    """
    name = name.replace(" ", "_").replace("/", "_")
    invalid = r'\:*?"<>|()'
    for ch in invalid:
        name = name.replace(ch, "")
    # Verwijder komma's en eventuele opeenvolgende underscores die achterblijven
    name = name.replace(",", "")
    while "__" in name:
        name = name.replace("__", "_")
    return name.strip("_")


def _parse_to_local(int_val: int, prefix_code: int) -> int:
    """The function `parse_to_local` removes a specified prefix code from an integer value."""
    str_val = str(int_val)
    if not str_val.startswith(f"{prefix_code}"):
        return int_val

    return int(str_val[len(str(prefix_code)) :])


def ParseList(val: str, prefix_code: int | None) -> list[int] | int:
    """The function `ParseList` checks if a given string represents a list and returns the list or the original value accordingly.

    Parameters
    ----------
    val: str
        The `ParseList` function takes a single parameter `val`, which is expected to be a string. The
    function checks if the string starts and ends with square brackets `[ ]`, indicating a list-like
    structure. If the string meets these conditions, it attempts to parse the string using
    `ast.literal_eval
    prefix_code : int | None
        The `prefix_code`, AQUO prefix. If specified it is to be removed from the integer values in the list.

    Returns
    -------
        The `ParseList` function is designed to parse a string representation of a list. If the input `val`
    is a string that starts with '[' and ends with ']', it attempts to evaluate the string using
    `ast.literal_eval` to convert it into a Python list. If the evaluation is successful and the result
    is a list, it returns the first element of the list if the list has only one element in it.

    """
    parsed = ast.literal_eval(val)

    if prefix_code is not None:
        parsed = [_parse_to_local(v, prefix_code) for v in parsed]

    # val = first item in list
    if len(parsed) == 1:
        return parsed[0]
    else:
        return parsed


def ReadOutputFile(
    model_folder,
    filetype,
    output_is_feather: bool = True,
    output_is_nc: bool = False,
    resample_to_daily: bool = True,
) -> pd.DataFrame:
    """Leest een Ribasim outputbestand en retourneert een DataFrame met daggemiddelden.

    Parameters
    ----------
    model_folder
        Map van het Ribasim-model; outputbestanden worden gelezen uit ``{model_folder}/results/``.
    filetype
        Type outputbestand. Mogelijke waarden: ``'basin'``, ``'flow'``, ``'control'``,
        ``'solver_stats'``, ``'basin_state'``.
    output_is_feather
        Lees het bestand als Arrow/Feather (``.arrow``). Standaard ``True``.
    output_is_nc
        Lees het bestand als NetCDF (``.nc``). Standaard ``False``.
    resample_to_daily
        Als ``True`` (standaard) wordt het tijdstap van de data gedetecteerd en automatisch
        naar daggemiddelden geresamplet als het interval korter dan één dag is. Zet op
        ``False`` als de data al dagwaarden bevat en je de hersampling wilt overslaan.

    Returns
    -------
    pd.DataFrame
        DataFrame met kolommen ``time``, ``link_id`` (of vergelijkbaar) en waarden,
        altijd op dagresolutie wanneer ``resample_to_daily=True``.
    """
    possible_filetypes = ["basin", "flow", "control", "solver_stats", "basin_state", "concentration"]
    if filetype.lower() not in possible_filetypes:
        raise ValueError(f"{filetype} not available. Choose one of the following: {possible_filetypes}")

    if filetype.lower() == "concentration" and output_is_feather:
        raise ValueError("'concentration' is alleen beschikbaar als NetCDF. Gebruik output_is_nc=True.")

    if output_is_feather and output_is_nc:
        raise ValueError("Both output_is_feather and output_is_nc are True. Choose only one.")
    if not output_is_feather and not output_is_nc:
        raise ValueError("Both output_is_feather and output_is_nc are False. Choose one format.")

    results_folder = Path(model_folder, "results")

    if output_is_feather:
        data = pd.read_feather(Path(results_folder, filetype.lower() + ".arrow"))
    else:
        ds = xr.open_dataset(Path(results_folder, filetype.lower() + ".nc"))
        data = ds.to_dataframe().reset_index()

    # Optional: fix timing if needed
    # data['time'] = data['time'] - pd.DateOffset(years=12)

    if resample_to_daily:
        data = _resample_to_daily(data)

    return data


def _resample_to_daily(data: pd.DataFrame) -> pd.DataFrame:
    """Resamplet een lang-formaat DataFrame naar daggemiddelden als het tijdstap sub-dagelijks is.

    Detecteert het mediane tijdstap en doet niets als de data al op dagresolutie staat.
    Werkt voor elk lang-formaat outputbestand: groepeert op alle niet-tijd/niet-waarde kolommen
    en middelt de numerieke waardekolommen per dag.
    """
    if "time" not in data.columns:
        return data

    data["time"] = pd.to_datetime(data["time"])
    unique_times = data["time"].drop_duplicates().sort_values()
    if len(unique_times) < 2:
        return data

    median_step = (unique_times.diff().dropna()).median()
    one_day = pd.Timedelta(days=1)

    if median_step >= one_day:
        # Al dagelijks of grover — niets te doen
        return data

    print(f"  Tijdstap gedetecteerd: {median_step}. Resampling naar daggemiddelden.")

    # Bepaal welke kolommen groepeersleutels zijn (niet-numeriek of identifier) en welke waarden
    id_cols = [
        c for c in data.columns if c != "time" and (data[c].dtype == object or not pd.api.types.is_float_dtype(data[c]))
    ]
    val_cols = [c for c in data.columns if c != "time" and c not in id_cols]

    data = data.groupby([*id_cols, pd.Grouper(key="time", freq="D")])[val_cols].mean().reset_index()

    return data


def LaadKoppeltabel(loc_koppeltabel, apply_for_water_authority: str | None = None) -> pd.DataFrame:
    """The function `LaadKoppeltabel` reads an Excel file, parses lists in the 'link_id' column, and converts the 'geometry' column to a geometry object.

    Parameters
    ----------
    loc_koppeltabel
        The `loc_koppeltabel` parameter in the `LaadKoppeltabel` function is expected to be a file location
    pointing to an Excel file that contains data for a koppeltabel (linking table).
    apply_for_water_authority
        Optional specification to read koppeltabel for a specific water authority. Defaults to None

    Returns
    -------
        The function returns the updated koppeltabel dataframe with the
    parsed columns 'link_id_parsed' and 'geometry_parsed'.

    """
    koppeltabel = pd.read_excel(loc_koppeltabel)

    # filter for water authority if specified
    if apply_for_water_authority is not None:
        koppeltabel = koppeltabel[koppeltabel["Waterschap"] == apply_for_water_authority]
        prefix_code = waterbeheercode[apply_for_water_authority]
    else:
        prefix_code = None

    # Convert the lists in link_id to lists if possible
    koppeltabel["link_id_parsed"] = koppeltabel["new_link_id"].apply(ParseList, args=(prefix_code,))

    # Parse the geometry
    koppeltabel["geometry_parsed"] = koppeltabel["geometry"].apply(lambda x: wkt.loads(x))

    return koppeltabel


def LaadSpecifiekeBewerking(loc_specifics) -> pd.DataFrame:
    """
    The function `LaadSpecifiekeBewerking` reads specific modifications from an Excel file and returns it.

    Parameters
    ----------
    loc_specifics:
        The `loc_specifics` parameter in the `LaadSpecifiekeBewerking` function is
        expected to be a file location pointing to an Excel file that contains specific modifications.

    Returns
    -------
    The function `LaadSpecifiekeBewerking` is returning the data read from an Excel file
    located at the path specified by the `loc_specifics` parameter as a pd.DataFrame
    """
    specifics = pd.read_excel(loc_specifics, header=0)

    return specifics


def get_unique(items):
    seen = []
    for item in items:
        if not any(item == s for s in seen):
            seen.append(item)
    return seen


def _safe_eval_formula(formula: str, env: dict) -> pd.Series:
    """Evaluate a simple arithmetic formula string safely using AST.

    Only allows numeric constants, variables present in `env`, and +/-/* / operators.
    """
    _bin_ops = {
        ast.Add: operator.add,
        ast.Sub: operator.sub,
        ast.Mult: operator.mul,
        ast.Div: operator.truediv,
    }
    _unary_ops = {
        ast.USub: operator.neg,
    }

    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Constant):
            if not isinstance(node.value, (int, float)):
                raise ValueError(f"Only numeric constants allowed: {node.value}")
            return node.value
        if isinstance(node, ast.Name):
            if node.id not in env:
                raise ValueError(f"Unknown variable: {node.id}")
            return env[node.id]
        if isinstance(node, ast.BinOp):
            op = _bin_ops.get(type(node.op))  # pyrefly: ignore[bad-argument-type]
            if op is None:
                raise ValueError(f"Unsupported operator: {type(node.op)}")
            return op(_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp):
            op = _unary_ops.get(type(node.op))  # pyrefly: ignore[bad-argument-type]
            if op is None:
                raise ValueError(f"Unsupported operator: {type(node.op)}")
            return op(_eval(node.operand))
        raise ValueError(f"Unsupported expression type: {type(node)}")

    return _eval(ast.parse(formula, mode="eval"))


def ApplySpecificOperation(data: pd.DataFrame, link: list | int, spec_op: str):
    """
    The function `ApplySpecificOperation`  performs specific operations on the input data.

    Parameters
    ----------
    data : pd.DataFrame
        The function `ApplySpecificOperation` takes in a pd.DataFrame `data`, a list of link IDs
    `link`, and a specific operation `spec_op` to be applied on the data. The function then performs
    different operations based on the value of `spec_op`.
    link : list | int
        The `link` parameter in the function `ApplySpecificOperation` is used to specify the link or links
    for which the specific operation will be applied. It can be either an integer or a list of integers
    representing the link IDs. If it is an integer, it will be converted to a list containing
    spec_op : str
        The `spec_op` parameter in the `ApplySpecificOperation` function represents a specific operation
    that can be performed on the data based on the provided conditions. The function uses a `match`
    statement to determine the specific operation to be applied. The possible values for `spec_op` and
    their corresponding operations are 'optellen','optellen_en_negatief_maken','negatief_maken', np.nan, or a
    custom formula. A custom formula needs to refer to the specific links as link1, link2 etc.

    Returns
    -------
    A subset of the data with a special operation applied (if any)
    """
    # Make sure the link is a list
    if isinstance(link, int):
        link = [link]

    # There are four options for the specific operation
    match spec_op:
        case "optellen":
            # Tel de links bij elkaar op wanneer de specifieke bewerking hierom vraagt
            subset_links = data[data["link_id"].isin(link)]
            # pyrefly: ignore[bad-assignment]
            subset_output: pd.DataFrame = subset_links.groupby("time", as_index=False)["flow_rate"].sum()

        case "negatief_maken":
            # Maak de meetreeks negatief
            subset_output = data[data["link_id"].isin(link)].copy()
            subset_output["flow_rate"] = subset_output["flow_rate"] * -1

        case "optellen_en_negatief_maken":
            # Tel op en maak de reeks negatief
            subset_links = data[data["link_id"].isin(link)]
            # pyrefly: ignore[bad-assignment]
            subset_output = subset_links.groupby("time", as_index=False)["flow_rate"].sum().copy()
            subset_output["flow_rate"] = subset_output["flow_rate"] * -1

        case _ if pd.isna(spec_op):
            # Als er geen specifieke bewerking nodig is, selecteer de link
            subset_output = data[data["link_id"].isin(link)]

        case _:
            # Handel de specifieke formule af
            link_mapping = {f"link{i + 1}": ID for i, ID in enumerate(link)}
            subset_links = data[data["link_id"].isin(link)]
            subset_pivot = subset_links.pivot(index="time", columns="link_id", values="flow_rate")
            env = {placeholder: subset_pivot[link_id] for placeholder, link_id in link_mapping.items()}
            subset_pivot["result"] = _safe_eval_formula(spec_op, env)
            subset_output = subset_pivot["result"].reset_index().rename(columns={"result": "flow_rate"})

    return subset_output


def AddCumulative(data, decade=False):
    """
    Calculates cumulative flow rates and sums for model output and measurements, respectively, with an option to adjust for decadal data.

    Parameters
    ----------
    data
        The `data` parameter is a pd.DataFrame containing the measurement data (column 'sum') and model data (column 'flow_rate')
    decade, optional
        The `decade` parameter is a boolean parameter that determines whether the data should be treated as decadal data. If `decade` is set to `True`, the function will
    multiply the cumulative values by 10 to represent decadal data.

    Returns
    -------
        The function `AddCumulative` returns a new dataset with two additional columns: "flow_rate_cum" and
    "sum_cum". The "flow_rate_cum" column contains the cumulative sum of the "flow_rate" data multiplied
    by a multiplier (which is 86400 by default, or 864000 if the `decade` parameter is set to True). The
    "sum_cum" column contains the cumulative sum of the measurements in the column 'sum'

    """
    # If decadal data is given, multiply by 10 to get the cumulative graph
    multiplier_s_to_d = 86400
    if decade:
        multiplier_s_to_d *= 10

    # Get the cumulative data for both the model output and the measurements
    new_data = data.copy()

    # Get cumulative of model output in
    new_data["flow_rate_cum"] = np.cumsum(new_data["flow_rate"]) * multiplier_s_to_d

    # Get cumulative data of measurements by filling the gaps
    new_data["sum_cum"] = new_data["sum"].fillna(0).cumsum() * multiplier_s_to_d

    return new_data


def GetStatisticsComparison(combined_df: pd.DataFrame) -> dict:
    """Berekent NSE, RMSE, MAE, relatieve bias, KGE en percentielen (P10/P25/P75/P90).

    Berekening op basis van de gecombineerde model- en meetreeks DataFrame.

    Parameters
    ----------
    combined_df
        DataFrame met kolommen 'flow_rate' (model) en 'sum' (meting).
        Rijen met NaN in 'sum' worden uitgesloten van de berekening.

    Returns
    -------
    dict met alle statistieken.
    """
    stats: dict[str, float | int] = {}

    # Verwijder tijdstappen zonder meting
    valid = combined_df.dropna(subset=["sum"])
    targets = np.asarray(valid["sum"], dtype=float)
    model_output = np.asarray(valid["flow_rate"], dtype=float)

    if len(targets) == 0:
        return dict.fromkeys(
            [
                "n_obs",
                "NSE",
                "RMSE",
                "MAE",
                "RelBias",
                "mean_sim",
                "mean_obs",
                "KGE",
                "P10_obs",
                "P10_sim",
                "P10_reldev",
                "P25_obs",
                "P25_sim",
                "P25_reldev",
                "P75_obs",
                "P75_sim",
                "P75_reldev",
                "P90_obs",
                "P90_sim",
                "P90_reldev",
            ],
            np.nan,
        )

    # --- Aantal geldige observaties ---
    stats["n_obs"] = len(targets)

    # --- NSE ---
    denom = np.sum((targets - np.mean(targets)) ** 2)
    stats["NSE"] = float(1 - np.sum((targets - model_output) ** 2) / denom) if denom != 0 else -999.0

    # --- RMSE ---
    stats["RMSE"] = float(np.sqrt(np.mean((model_output - targets) ** 2)))

    # --- MAE ---
    stats["MAE"] = float(np.mean(np.abs(model_output - targets)))

    # --- Relatieve bias op het gemiddeld debiet [%] ---
    mean_obs = np.mean(targets)
    if mean_obs != 0:
        stats["RelBias"] = float((np.mean(model_output) - mean_obs) / mean_obs * 100)
    else:
        stats["RelBias"] = np.nan
    stats["mean_sim"] = float(np.mean(model_output))
    stats["mean_obs"] = float(mean_obs)

    # --- KGE (Kling-Gupta Efficiency) ---
    r = float(np.corrcoef(targets, model_output)[0, 1]) if np.std(targets) > 0 and np.std(model_output) > 0 else np.nan
    alpha = float(np.std(model_output) / np.std(targets)) if np.std(targets) != 0 else np.nan
    beta = float(np.mean(model_output) / mean_obs) if mean_obs != 0 else np.nan
    if any(np.isnan(v) for v in [r, alpha, beta]):
        stats["KGE"] = np.nan
    else:
        stats["KGE"] = float(1 - np.sqrt((r - 1) ** 2 + (alpha - 1) ** 2 + (beta - 1) ** 2))

    # --- Percentielen P10, P25, P75, P90 ---
    for p in [10, 25, 75, 90]:
        obs_p = float(np.nanpercentile(targets, p))
        sim_p = float(np.nanpercentile(model_output, p))
        stats[f"P{p}_obs"] = obs_p
        stats[f"P{p}_sim"] = sim_p
        stats[f"P{p}_reldev"] = float((sim_p - obs_p) / abs(obs_p) * 100) if obs_p != 0 else np.nan

    return stats


# Interne kleurentabel voor Excel-opmaak (PatternFill-objecten worden eenmalig aangemaakt)
_BEOOR_KLEUREN = {
    "Goed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Matig": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Onvoldoende": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "n.v.t.": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


def BeoordeelCriteria(stats: dict, criteria_grenzen: dict | None = None) -> dict:
    """Beoordeelt de statistieken op basis van de KRW-toetsingscriteria.

    Parameters
    ----------
    stats
        Dictionary met statistieken zoals teruggegeven door GetStatisticsComparison.
    criteria_grenzen
        Grenswaarden per criterium. Standaard KRW-drempelwaarden worden gebruikt als
        None wordt meegegeven.

    Returns
    -------
    dict met per criterium de beoordeling: 'Goed', 'Matig', 'Onvoldoende' of 'n.v.t.'
    """
    if criteria_grenzen is None:
        criteria_grenzen = {
            "Bias": {"Goed": 10.0, "Matig": 15.0},
            "NSE": {"Goed": 0.5},
            "KGE": {"Goed": 0.6},
            "P10": {"Goed": 25.0, "Matig": 30.0},
            "P25": {"Goed": 10.0, "Matig": 15.0},
            "P75": {"Goed": 10.0, "Matig": 15.0},
            "P90": {"Goed": 25.0, "Matig": 30.0},
        }

    beoordeling = {}

    def _beoordeel(waarde, criterium, hoger_is_beter=False):
        if waarde is None or np.isnan(waarde):
            return "n.v.t."
        grenzen = criteria_grenzen[criterium]
        if hoger_is_beter:
            return "Goed" if waarde > grenzen["Goed"] else "Onvoldoende"
        else:
            abs_val = abs(waarde)
            if "Matig" in grenzen:
                return (
                    "Goed" if abs_val < grenzen["Goed"] else ("Matig" if abs_val < grenzen["Matig"] else "Onvoldoende")
                )
            else:
                return "Goed" if abs_val < grenzen["Goed"] else "Onvoldoende"

    beoordeling["Bias"] = _beoordeel(stats.get("RelBias"), "Bias")
    beoordeling["NSE"] = _beoordeel(stats.get("NSE"), "NSE", hoger_is_beter=True)
    beoordeling["KGE"] = _beoordeel(stats.get("KGE"), "KGE", hoger_is_beter=True)
    beoordeling["P10"] = _beoordeel(stats.get("P10_reldev"), "P10")
    beoordeling["P25"] = _beoordeel(stats.get("P25_reldev"), "P25")
    beoordeling["P75"] = _beoordeel(stats.get("P75_reldev"), "P75")
    beoordeling["P90"] = _beoordeel(stats.get("P90_reldev"), "P90")

    return beoordeling


def GetStatisticsPerPeriod(combined_df: pd.DataFrame) -> dict:
    """Berekent statistieken voor elk uniek jaar in de data én voor de totale periode.

    De perioden worden afgeleid vanuit de Ribasim modeluitvoer zelf, zodat het script
    automatisch werkt voor andere tijdsperioden dan 2017-2019.

    Parameters
    ----------
    combined_df
        DataFrame met kolommen 'time', 'flow_rate' en 'sum'.

    Returns
    -------
    dict met per jaar (als string) en 'totaal' de statistieken.
    """
    results: dict[str, dict | None] = {}
    unique_years = sorted(combined_df["time"].dt.year.unique())

    for year in unique_years:
        df_year = combined_df[combined_df["time"].dt.year == year]
        if df_year.dropna(subset=["sum"]).shape[0] >= 5:
            results[str(year)] = GetStatisticsComparison(df_year)
        else:
            results[str(year)] = None  # te weinig geldige metingen

    results["totaal"] = GetStatisticsComparison(combined_df)
    return results


def SaveTimeseries(
    combined_df: pd.DataFrame,
    combined_df_decade: pd.DataFrame,
    fig_name_clean: str,
    meetreeks_naam: str,
    output_folder: str,
) -> None:
    """Schrijft de tijdsreeksen (dag en decade) weg als Excel-bestand met twee sheets.

    Parameters
    ----------
    combined_df
        DataFrame met dagwaarden (kolommen: time, flow_rate, sum).
    combined_df_decade
        DataFrame met decadewaarden (kolommen: time, flow_rate, sum).
    fig_name_clean
        Bestandsnaam (zonder extensie) afgeleid van de meetreeksnaam.
    meetreeks_naam
        Naam van de meetreeks, gebruikt als kolomnaam in de output.
    output_folder
        Map waar het Excel-bestand wordt opgeslagen.
    """
    Path(output_folder).mkdir(parents=True, exist_ok=True)
    filepath = Path(output_folder) / f"{fig_name_clean}.xlsx"

    # Dag sheet: datum | meetreeks | Ribasim
    dag_df = combined_df[["time", "sum", "flow_rate"]].copy()
    dag_df.columns = ["Datum", meetreeks_naam, "Ribasim [m3/s]"]

    # Decade sheet: datum | meetreeks | Ribasim
    dec_df = combined_df_decade[["time", "sum", "flow_rate"]].copy()
    dec_df.columns = ["Datum", meetreeks_naam, "Ribasim [m3/s]"]

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        dag_df.to_excel(writer, sheet_name="Dag", index=False)
        dec_df.to_excel(writer, sheet_name="Decade", index=False)


def _KleureerBeoordelingen(ws, df: pd.DataFrame, beoor_kleuren: dict | None = None) -> None:
    """Kleurt de beoordelingskolommen (kolommen die 'Beoor_' bevatten) in het worksheet.

    Parameters
    ----------
    ws
        openpyxl worksheet object.
    df
        DataFrame waaruit het worksheet is aangemaakt (voor kolomindexen).
    beoor_kleuren
        Kleurentabel voor Excel-opmaak. Standaard: interne ``_BEOOR_KLEUREN``.
    """
    if beoor_kleuren is None:
        beoor_kleuren = _BEOOR_KLEUREN
    beoor_cols = [i + 1 for i, col in enumerate(df.columns) if col.startswith("Beoor_")]
    for row_idx in range(2, ws.max_row + 1):  # rij 1 = header
        for col_idx in beoor_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            kleur = beoor_kleuren.get(cell.value, beoor_kleuren["n.v.t."])
            cell.fill = kleur


def BerekenModelEindbeoordeling(
    validatie_xlsx: str | Path,
    output_xlsx: str | Path | None = None,
    periode: str = "totaal",
    criteria_groepen: dict | None = None,
    min_groepen_voldoen: int = 2,
    drempel_hws: float = 75.0,
    drempel_regionaal: float = 50.0,
    hws_waterschap: str = "Rijkswaterstaat",
    beoor_kleuren: dict | None = None,
) -> pd.DataFrame:
    """Leest de weggeschreven Validatie_criteria.xlsx en berekent de eindbeoordeling.

    Werkt op de decade-beoordelingskolommen (Beoor_*_dec) die per waterschap-sheet
    aanwezig zijn. De criteria zijn gegroepeerd in drie overkoepelende groepen
    (``criteria_groepen``):

    - Bias         : RelBias_dec
    - NSE/KGE      : NSE_dec en/of KGE_dec  (Goed als ≥ 1 Goed)
    - Percentielen : P10, P25, P75, P90      (Goed als ≥ 2 van 4 Goed)

    Een locatie 'voldoet' als ``min_groepen_voldoen`` (standaard 2) van de 3 groepen Goed zijn.
    Het model is 'Geschikt' als het % voldoende locaties de drempel haalt:
    - ``drempel_hws`` voor sheet ``hws_waterschap`` (hoofdwatersysteem)
    - ``drempel_regionaal`` voor overige sheets

    Parameters
    ----------
    validatie_xlsx
        Pad naar de weggeschreven Validatie_criteria.xlsx.
    output_xlsx
        Optioneel uitvoerpad voor een nieuwe Excel met de eindbeoordeling.
    periode
        Rij-filter op de kolom 'Periode'; standaard 'totaal'.
    criteria_groepen
        Groepdefinities met kolommen en min_goed per groep.
    min_groepen_voldoen
        Minimaal aantal groepen dat 'Goed' moet zijn voor een locatie om te voldoen.
    drempel_hws
        Minimaal % voldoende locaties voor het hoofdwatersysteem.
    drempel_regionaal
        Minimaal % voldoende locaties voor regionale waterschappen.
    hws_waterschap
        Naam van het HWS-sheet (krijgt strengere drempel).
    beoor_kleuren
        Kleurentabel voor Excel-opmaak. Standaard: interne ``_BEOOR_KLEUREN``.

    Returns
    -------
    pd.DataFrame met de eindbeoordeling per waterschap + samenvattingsrijen.
    """
    if criteria_groepen is None:
        criteria_groepen = {
            "Bias": {"kolommen": ["Beoor_Bias_dec"], "min_goed": 1},
            "NSE_KGE": {"kolommen": ["Beoor_NSE_dec", "Beoor_KGE_dec"], "min_goed": 1},
            "Percentielen": {
                "kolommen": ["Beoor_P10_dec", "Beoor_P25_dec", "Beoor_P75_dec", "Beoor_P90_dec"],
                "min_goed": 4,
            },
        }
    if beoor_kleuren is None:
        beoor_kleuren = _BEOOR_KLEUREN

    # Open de validatie-Excel zonder alle sheets tegelijk in geheugen te laden
    xl = pd.ExcelFile(validatie_xlsx)
    detail_rijen: list[dict] = []
    samenvatting_rijen: list[dict] = []

    # Vertaaltabel van beoordelingswaarden naar Excel-opvulkleuren
    _kleur_map = {
        "Geschikt": beoor_kleuren["Goed"],
        "Ongeschikt": beoor_kleuren["Onvoldoende"],
        "Ja": beoor_kleuren["Goed"],
        "Nee": beoor_kleuren["Onvoldoende"],
        "Goed": beoor_kleuren["Goed"],
        "Onvoldoende": beoor_kleuren["Onvoldoende"],
    }

    # ── Stap 1: doorloop elke waterschap-sheet ────────────────────────────────
    for sheet in xl.sheet_names:
        df = xl.parse(sheet)
        # Sla sheets over die geen statistieken-structuur hebben (bijv. extra info-sheets)
        if "Periode" not in df.columns:
            continue
        # Filter op de gewenste periode (standaard "totaal" = statistieken over de gehele simulatieperiode)
        df_p = df[df["Periode"] == periode].copy()
        if df_p.empty:
            continue

        # Rijkswaterstaat krijgt een strengere drempel dan regionale waterschappen
        is_hws = sheet == hws_waterschap
        drempel = drempel_hws if is_hws else drempel_regionaal
        n_voldoet = 0

        # ── Stap 2: beoordeel elke locatie ───────────────────────────────────
        for _, rij in df_p.iterrows():
            groep_beoor: dict[str, str] = {}
            for groep, cfg in criteria_groepen.items():
                # Alleen kolommen meenemen die ook echt in de sheet aanwezig zijn
                aanwezig = [k for k in cfg["kolommen"] if k in df_p.columns]
                n_goed = sum(rij.get(k, "n.v.t.") == "Goed" for k in aanwezig)
                # Groep is "Goed" als het minimum aantal "Goed"-kolommen is gehaald
                groep_beoor[groep] = "Goed" if (aanwezig and n_goed >= cfg["min_goed"]) else "Onvoldoende"

            # Locatie voldoet als minimaal min_groepen_voldoen van de 3 groepen "Goed" zijn
            n_groepen_goed = sum(v == "Goed" for v in groep_beoor.values())
            voldoet = n_groepen_goed >= min_groepen_voldoen
            if voldoet:
                n_voldoet += 1

            # Sla het locatie-detailresultaat op voor de Details_per_locatie-sheet
            detail_rijen.append(
                {
                    "Waterschap": sheet,
                    "Locatie": rij.get("Locatie", "onbekend"),
                    **{f"Groep_{g}": v for g, v in groep_beoor.items()},
                    "N_groepen_Goed": n_groepen_goed,
                    "Voldoet": "Ja" if voldoet else "Nee",
                }
            )

        # ── Stap 3: samenvatting per waterschap ──────────────────────────────
        n_totaal = len(df_p)
        pct = round(n_voldoet / n_totaal * 100, 1) if n_totaal > 0 else np.nan
        samenvatting_rijen.append(
            {
                "Waterschap": sheet,
                "Type": "HWS" if is_hws else "Regionaal",
                "N_locaties": n_totaal,
                "N_voldoen": n_voldoet,
                "Pct_voldoen": pct,
                "Drempel [%]": drempel,
                # Model "Geschikt" als het percentage voldoende locaties de drempel haalt
                "Beoordeling": "Geschikt" if (not np.isnan(pct) and pct >= drempel) else "Ongeschikt",
            }
        )

    if not samenvatting_rijen:
        print("Geen bruikbare sheets gevonden in de validatie-Excel.")
        return pd.DataFrame()

    df_sam = pd.DataFrame(samenvatting_rijen)
    df_det = pd.DataFrame(detail_rijen)

    # ── Stap 4: voeg totaalrijen toe voor HWS en Regionaal als geheel ────────
    for type_naam, type_drempel in [("HWS", drempel_hws), ("Regionaal", drempel_regionaal)]:
        sub = df_sam[df_sam["Type"] == type_naam]
        if sub.empty:
            continue
        n_tot = int(sub["N_locaties"].sum())
        n_vol = int(sub["N_voldoen"].sum())
        pct = round(n_vol / n_tot * 100, 1) if n_tot > 0 else np.nan
        df_sam = pd.concat(
            [
                df_sam,
                pd.DataFrame(
                    [
                        {
                            "Waterschap": f"── Totaal {type_naam}",
                            "Type": type_naam,
                            "N_locaties": n_tot,
                            "N_voldoen": n_vol,
                            "Pct_voldoen": pct,
                            "Drempel [%]": type_drempel,
                            "Beoordeling": "Geschikt" if (not np.isnan(pct) and pct >= type_drempel) else "Ongeschikt",
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )

    # ── Stap 5: wegschrijven naar Excel (optioneel) ───────────────────────────
    if output_xlsx is not None:
        Path(output_xlsx).resolve().parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
            # Sheet 1: samenvatting per waterschap met eindbeoordeling
            df_sam.to_excel(writer, sheet_name="Eindbeoordeling", index=False)
            ws = writer.sheets["Eindbeoordeling"]
            # Kleureer de Beoordeling-kolom (Geschikt = groen, Ongeschikt = rood)
            beoor_col = df_sam.columns.tolist().index("Beoordeling") + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=beoor_col)
                if cell.value in _kleur_map:
                    cell.fill = _kleur_map[cell.value]
            # Notitieregel met de gebruikte instellingen voor reproduceerbaarheid
            ws.cell(row=ws.max_row + 2, column=1).value = (
                f"Instellingen: periode={periode}  |  tijdreeks=decade  |  "
                f"min. groepen voldoen={min_groepen_voldoen}/3  |  "
                f"drempel HWS={drempel_hws}%  |  drempel regionaal={drempel_regionaal}%"
            )

            # Sheet 2: per locatie de drie groepsbeoordelingen en of de locatie voldoet
            if not df_det.empty:
                df_det.to_excel(writer, sheet_name="Details_per_locatie", index=False)
                ws_det = writer.sheets["Details_per_locatie"]
                # Kleureer de Groep_*- en Voldoet-kolommen
                kleur_cols = [i + 1 for i, c in enumerate(df_det.columns) if c.startswith("Groep_") or c == "Voldoet"]
                for row_idx in range(2, ws_det.max_row + 1):
                    for col_idx in kleur_cols:
                        cell = ws_det.cell(row=row_idx, column=col_idx)
                        if cell.value in _kleur_map:
                            cell.fill = _kleur_map[cell.value]

        print(f"Eindbeoordeling opgeslagen: {output_xlsx}")

    return df_sam


def ExportToExcel(
    results_excel: dict, output_path: str | Path, criteria_grenzen: dict | None, beoor_kleuren: dict | None = None
) -> None:
    """Schrijft per waterschap een sheet naar een Excel-bestand.

    Bevat alle statistieken en beoordelingen voor dag- en decadewaarden,
    uitgesplitst per periode.

    Parameters
    ----------
    results_excel
        Dict met structuur:
        {waterschap: [{"locatie": str,
                       "stats_dag":  {periode: stats_dict | None},
                       "stats_dec":  {periode: stats_dict | None}}, ...]}
    output_path
        Pad naar het te schrijven Excel-bestand.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for waterschap, records in results_excel.items():
            rijen = []
            for record in records:
                perioden = [*sorted(p for p in record["stats_dag"] if p != "totaal"), "totaal"]

                for periode in perioden:
                    stats_d = record["stats_dag"].get(periode)
                    stats_dc = record["stats_dec"].get(periode)

                    def _fmt(stats, key):
                        if stats is None or stats.get(key) is None:
                            return np.nan
                        val = stats[key]
                        return round(float(val), 3) if not np.isnan(val) else np.nan

                    beoor_d = (
                        BeoordeelCriteria(stats_d, criteria_grenzen)
                        if stats_d is not None
                        else dict.fromkeys(["NSE", "KGE", "Bias", "P10", "P25", "P75", "P90"], "n.v.t.")
                    )
                    beoor_dc = (
                        BeoordeelCriteria(stats_dc, criteria_grenzen)
                        if stats_dc is not None
                        else dict.fromkeys(["NSE", "KGE", "Bias", "P10", "P25", "P75", "P90"], "n.v.t.")
                    )

                    rij = {
                        "Locatie": record["locatie"],
                        "Periode": periode,
                        # --- Dagwaarden: statistieken ---
                        "n_obs_dag": _fmt(stats_d, "n_obs"),
                        "NSE_dag": _fmt(stats_d, "NSE"),
                        "KGE_dag": _fmt(stats_d, "KGE"),
                        "RelBias_dag [%]": _fmt(stats_d, "RelBias"),
                        "P10_reldev_dag [%]": _fmt(stats_d, "P10_reldev"),
                        "P25_reldev_dag [%]": _fmt(stats_d, "P25_reldev"),
                        "P75_reldev_dag [%]": _fmt(stats_d, "P75_reldev"),
                        "P90_reldev_dag [%]": _fmt(stats_d, "P90_reldev"),
                        # --- Dagwaarden: beoordelingen ---
                        "Beoor_NSE_dag": beoor_d["NSE"],
                        "Beoor_KGE_dag": beoor_d["KGE"],
                        "Beoor_Bias_dag": beoor_d["Bias"],
                        "Beoor_P10_dag": beoor_d["P10"],
                        "Beoor_P25_dag": beoor_d["P25"],
                        "Beoor_P75_dag": beoor_d["P75"],
                        "Beoor_P90_dag": beoor_d["P90"],
                        # --- Decadewaarden: statistieken ---
                        "n_obs_dec": _fmt(stats_dc, "n_obs"),
                        "NSE_dec": _fmt(stats_dc, "NSE"),
                        "KGE_dec": _fmt(stats_dc, "KGE"),
                        "RelBias_dec [%]": _fmt(stats_dc, "RelBias"),
                        "P10_reldev_dec [%]": _fmt(stats_dc, "P10_reldev"),
                        "P25_reldev_dec [%]": _fmt(stats_dc, "P25_reldev"),
                        "P75_reldev_dec [%]": _fmt(stats_dc, "P75_reldev"),
                        "P90_reldev_dec [%]": _fmt(stats_dc, "P90_reldev"),
                        # --- Decadewaarden: beoordelingen ---
                        "Beoor_NSE_dec": beoor_dc["NSE"],
                        "Beoor_KGE_dec": beoor_dc["KGE"],
                        "Beoor_Bias_dec": beoor_dc["Bias"],
                        "Beoor_P10_dec": beoor_dc["P10"],
                        "Beoor_P25_dec": beoor_dc["P25"],
                        "Beoor_P75_dec": beoor_dc["P75"],
                        "Beoor_P90_dec": beoor_dc["P90"],
                    }
                    rijen.append(rij)

            if not rijen:
                continue

            df_sheet = pd.DataFrame(rijen)
            sheet_name = waterschap[:31]  # Excel beperkt sheetnamen tot 31 tekens
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

            # Kleurcodering op beoordelingskolommen
            ws = writer.sheets[sheet_name]
            _KleureerBeoordelingen(ws, df_sheet, beoor_kleuren=beoor_kleuren)

            # Kolombreedte automatisch aanpassen
            for col_idx, col_name in enumerate(df_sheet.columns, start=1):
                max_len = max(len(str(col_name)), df_sheet[col_name].astype(str).str.len().max())
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 30)


def _PlotAndSaveFractie(
    combined_df: pd.DataFrame,
    concentration_path: Path | None,
    basin_node_id: int,
    tracers: list[str] | None,
    fig_name: str,
    waterschap: str,
    output_folder: Path,
) -> str | None:
    """Gecombineerd figuur: gestapelde fracties (boven) + Ribasim+meting tijdreeks (onder).

    De twee panelen delen de x-as zodat pieken direct te correleren zijn met de
    bijdragende fracties. Opgeslagen in {output_folder}/{waterschap}/{fig_name}_fractie.png.

    Returns HTML img-tag string voor gebruik in popup, of None als er geen fractie
    data beschikbaar is voor dit basin.
    """
    if concentration_path is None:
        return None
    ds = xr.open_dataset(concentration_path)
    table = ds.to_dataframe().reset_index()
    ds.close()

    table = table[table["node_id"] == basin_node_id]
    if tracers is not None:
        table = table[table["substance"].isin(tracers)]

    if len(table) == 0:
        return None

    groups = table.groupby("substance")
    available = [t for t in (tracers or sorted(table["substance"].unique())) if t in groups.groups]
    stack = {k: groups.get_group(k)["concentration"].to_numpy() for k in available}
    key = next(iter(stack))
    frac_time = groups.get_group(key)["time"]

    font = "Arial"
    _halo = [pe.withStroke(linewidth=5, foreground="black")]
    fig, ax = plt.subplots(figsize=(11, 6))

    # Gestapelde fracties op linker y-as
    ax.stackplot(frac_time, list(stack.values()), labels=list(stack.keys()), alpha=0.75)
    ax.plot(frac_time, np.sum(list(stack.values()), axis=0), c="black", lw=1.8, label="_totaal")
    ax.set_ylabel("Fractie [-]", fontdict={"fontsize": 13, "fontname": font, "fontweight": "bold"})
    ax.set_ylim(0, 1.05)
    ax.grid(True, color="#555555", alpha=0.5, linewidth=0.7)
    ax.set_title(f"{fig_name} — Basin {basin_node_id}", fontsize=11, fontweight="bold", wrap=True)
    ax.tick_params(axis="both", labelsize=11)
    for lbl in ax.get_xticklabels() + ax.get_yticklabels():
        lbl.set_fontweight("bold")
    plt.setp(ax.get_xticklabels(), rotation=30, ha="right")

    # Debiet lijnen op rechter y-as; witte halo voorkomt verwarring met fractie-kleuren
    ax2 = ax.twinx()
    ax2.plot(
        combined_df["time"],
        combined_df["flow_rate"],
        label="Ribasim",
        color="tab:blue",
        linewidth=2.5,
        zorder=5,
        path_effects=_halo,
    )
    if "sum" in combined_df.columns and not combined_df["sum"].isna().all():
        ax2.plot(
            combined_df["time"],
            combined_df["sum"],
            label="Meting",
            color="tab:orange",
            linewidth=2.8,
            zorder=5,
            path_effects=_halo,
        )
    ax2.set_ylabel("Debiet [m$^3$/s]", fontdict={"fontsize": 13, "fontname": font, "fontweight": "bold"})
    ax2.tick_params(axis="y", labelsize=11)
    for lbl in ax2.get_yticklabels():
        lbl.set_fontweight("bold")

    # Gecombineerde legenda buiten de plotruimte (onder), zodat overlap met data
    # structureel onmogelijk is — twinx-lijnen worden niet meegenomen door loc="best"
    handles1, labels1 = ax.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(
        handles1 + handles2,
        labels1 + labels2,
        loc="upper center",
        bbox_to_anchor=(0.5, -0.20),
        ncol=8,
        prop={"family": font, "weight": "bold", "size": 10},
        frameon=True,
    )
    fig.subplots_adjust(bottom=0.20)

    fig_name_clean = _sanitize_filename(fig_name)
    fig_folder = output_folder / waterschap
    fig_folder.mkdir(parents=True, exist_ok=True)
    fname = fig_name_clean + "_fractie.png"
    fig.savefig(fig_folder / fname, bbox_inches="tight", dpi=150)
    plt.close(fig)

    return f'<img src="../figures_fracties/{waterschap}/{fname}" width=400 height=300>'


def CompareOutputMeasurements(
    loc_koppeltabel,
    loc_specifics,
    meas_folder,
    model_folder,
    toml_file: str,
    filetype="flow",
    criteria_grenzen: dict | None = None,
    beoor_kleuren: dict | None = None,
    apply_for_water_authority: str | None = None,
    save_results_combined: bool = False,
    output_is_feather: bool = True,
    output_is_nc: bool = False,
    resample_to_daily: bool = True,
) -> None:
    """Compares model output measurements with actual measurements, calculates statistics, and saves the results in a geopackage per waterboard as well as producing the necessary figures.

    Parameters
    ----------
    loc_koppeltabel
        The location of the koppeltabel (Excel file):
    loc_specifics
        The location of the table with specific operations per measurement (Excel file)
    meas_folder
        The `meas_folder` refers to the folder location where measurements data is stored.
    model_folder
        The `model_folder` refers to the directory path where the model is stored.
    filetype
        The `filetype` parameter specifies the type of output file to be read from the `model_folder`.
        By default, it is set to `'flow'`, but you can change it to
    apply_for_water_authority
        Optional specification to read koppeltabel for a specific water authority. Defaults to None
    save_results_combined
        Optional boolean to check if all the results are to be written to a single layer in a geopackage instead of
        one per water authority. Defaults to False
    model
        Optioneel geladen ribasim.Model object voor het genereren van gecombineerde
        fractie+tijdreeks plots. Vereist dat concentration.nc aanwezig is in
        model.results_path. Locaties met meerdere links of zonder bereikbaar Basin
        worden automatisch overgeslagen. Als None (standaard) worden geen fractie
        plots aangemaakt.

    Returns
    -------
    The function returns nothing, but saves the results in .png figures, geopackages,
    tijdsreeksen per meetlocatie en een Excel-bestand met criteria en beoordelingen.
    Als model is opgegeven en concentration.nc beschikbaar is, wordt ook
    Fractie_locaties.gpkg geschreven.

    """
    koppeltabel = LaadKoppeltabel(loc_koppeltabel, apply_for_water_authority=apply_for_water_authority)
    specifics = LaadSpecifiekeBewerking(loc_specifics)
    data = ReadOutputFile(
        model_folder,
        filetype,
        output_is_feather=output_is_feather,
        output_is_nc=output_is_nc,
        resample_to_daily=resample_to_daily,
    )

    measurements = LoadMeasurements(meas_folder)

    # ── Fractie plots setup (optioneel, vereist model met concentration.nc) ──────
    print("Inlezen model...")
    model = Model.read(Path(model_folder) / toml_file)

    _frac_available = False
    _concentration_path = None
    _tracers = None
    _frac_fig_folder = Path(model_folder) / "results" / "figures_fracties"
    if model is not None:
        try:
            _concentration_path = Path(model.results_path) / "concentration.nc"
            _frac_available = _concentration_path.exists()
            if not _frac_available:
                print(f"concentration.nc niet gevonden: {_concentration_path}. Fractie plots overgeslagen.")
        except AttributeError:
            print("model.results_path niet beschikbaar. Fractie plots overgeslagen.")
        if _frac_available and get_lhm_fractions is not None:
            try:
                _tracers = get_lhm_fractions(model)
                print(f"LHM tracers geladen ({len(_tracers)}): {_tracers}")
            except Exception as exc:
                print(f"get_lhm_fractions mislukt ({exc}). Standaard tracers worden gebruikt.")

    results_measurements: dict[str, dict[str, list[object]]] = {}
    results_measurements_decade: dict[str, dict[str, list[object]]] = {}
    results_excel = {}  # voor het Excel-criteriabestand

    # Aparte lijst voor fractie locaties (los van de validatie geopackages)
    _fractie_records: list[dict] = []

    # Pre-bouw graph voor O(1) basin-traversal in de loop hieronder
    _model_graph: dict | None = _build_model_graph(model) if model is not None else None

    # Get the unique link ids
    unique_links = get_unique(koppeltabel["link_id_parsed"])

    for link in tqdm.tqdm(unique_links, total=len(unique_links), desc="Verwerken metingen"):
        try:
            if np.isnan(link):
                print("A link with type NaN has been found. Skipped.")
                continue
        except:  # noqa: E722 S110 TODO: do not use bare except
            pass

        # for n, meetlocatie in tqdm.tqdm(koppeltabel.iterrows(), total=len(koppeltabel), desc='Verwerken metingen'):
        mask = koppeltabel["link_id_parsed"].apply(lambda x, _link=link: x == _link)
        meetlocaties_link = koppeltabel[mask]
        waterschap = meetlocaties_link.iloc[0]["Waterschap"]

        # Create a result dictionary per waterschap
        if waterschap not in results_measurements:

            def _leeg_results():
                return {
                    "koppelinfo": [],
                    "waterschap": [],
                    "link_id": [],
                    "MeetreeksC": [],
                    "Aan/Af": [],
                    "NSE": [],
                    "RMSE": [],
                    "MAE": [],
                    "RelBias": [],
                    "KGE": [],
                    "P10_reldev": [],
                    "P25_reldev": [],
                    "P75_reldev": [],
                    "P90_reldev": [],
                    "geometry": [],
                    "figure_path": [],
                    "QGIS_map_tip": [],
                }

            results_measurements[waterschap] = _leeg_results()
            results_measurements_decade[waterschap] = _leeg_results()
            results_excel[waterschap] = []

        # Loop over the locations in case two need to be summed
        # Two locs cannot be summed if the type of discharge differs
        unieke_aanaf = np.unique(meetlocaties_link["Aan/Af"])
        if len(unieke_aanaf) > 1:
            # HARDCODED uitzondering: AmstelGooienVecht mag metingen met hetzelfde
            # Aan/Af-type optellen ook als de link meerdere Aan/Af-waarden heeft.
            # Rijen met een afwijkend Aan/Af worden weggelaten; de rest wordt gesommeerd.
            if waterschap == "AmstelGooienVecht":
                meest_voorkomend = meetlocaties_link["Aan/Af"].mode()[0]
                meetlocaties_link = meetlocaties_link[meetlocaties_link["Aan/Af"] == meest_voorkomend]
                print(
                    f"[HARDCODED] Link {link} ({waterschap}): meerdere Aan/Af-waarden gevonden "
                    f"({list(unieke_aanaf)}). Rijen met Aan/Af='{meest_voorkomend}' worden opgeteld, "
                    f"overige rijen weggelaten."
                )
            else:
                dominant_aanaf = meetlocaties_link.iloc[0]["Aan/Af"]
                details = ", ".join(
                    f"{row['MeetreeksC']} ({row['Aan/Af']})"
                    for _, row in meetlocaties_link[["MeetreeksC", "Aan/Af"]].iterrows()
                )
                weggelaten = [
                    row["MeetreeksC"] for _, row in meetlocaties_link.iterrows() if row["Aan/Af"] != dominant_aanaf
                ]
                print(
                    f"Link {link} ({waterschap}): meerdere Aan/Af-waarden gevonden "
                    f"({list(unieke_aanaf)}). Doorgaan met Aan/Af='{dominant_aanaf}'. "
                    f"Meetreeksen: {details}." + (f" Weggelaten (andere Aan/Af): {weggelaten}." if weggelaten else "")
                )
                meetlocaties_link = meetlocaties_link[meetlocaties_link["Aan/Af"] == dominant_aanaf]

        if meetlocaties_link.iloc[0]["Aan/Af"] == "Aanvoer":
            dagmetingen = measurements["aanvoer_dag"]
        elif (meetlocaties_link.iloc[0]["Aan/Af"] == "Afvoer") or (meetlocaties_link.iloc[0]["Aan/Af"] == "Aan&Af"):
            dagmetingen = measurements["afvoer_dag"]
        else:
            print(f"Onbekend Aan/Af-type '{meetlocaties_link.iloc[0]['Aan/Af']}' voor link {link}, overgeslagen.")
            continue

        existing_measurements = [col for col in meetlocaties_link["MeetreeksC"] if col in dagmetingen.columns]
        missing_measurements = [col for col in meetlocaties_link["MeetreeksC"] if col not in dagmetingen.columns]
        if len(missing_measurements) > 0:
            for col in missing_measurements:
                print(f"Cannot find the daily measurements for {col}")

        subset_measurements = dagmetingen[["time", *existing_measurements]].copy()

        # If multiple measurements series refer to the same link, take the sum of the measurements
        subset_measurements["sum"] = subset_measurements[existing_measurements].sum(axis=1, min_count=1)

        # Get the specific operations for the measurements
        subset_specs = specifics[
            (specifics["MeetreeksC"].isin(existing_measurements))
            & (specifics["Aan/Af"] == meetlocaties_link.iloc[0]["Aan/Af"])
        ]

        if len(pd.Series.unique(subset_specs["Specifiek"])) > 1:
            print(
                f"Two different operations found for the same set of links. A.o. in measurements {existing_measurements}"
            )
            continue

        # Set the specific operation
        spec_op = subset_specs["Specifiek"].iloc[0]

        # If a list of links is present, a specific operation is required.
        if isinstance(link, list) & pd.isna(spec_op):
            print(f"No specific operation found for measurements {existing_measurements}, around link {link}")
            continue

        # Apply the special operation to get the subset of model output
        subset_modeloutput = ApplySpecificOperation(data, link, spec_op)

        # Combine the measurements with the modeloutput in one dataframe
        combined_df = subset_modeloutput.merge(subset_measurements[["time", "sum"]], on=["time"], how="left")

        # Check whether there are any measurements at all during the period
        if combined_df["sum"].isna().all().all():
            print(
                f"No measured data available for the requested time period for link {link}, a.o. location {meetlocaties_link.iloc[0]['MeetreeksC']}"
            )
            continue

        combined_df_decade = ConvertToDecade(combined_df)

        # Bereken statistieken voor totale periode en per jaar (afgeleid uit de data)
        stats_dag_per_periode = GetStatisticsPerPeriod(combined_df)
        stats_dec_per_periode = GetStatisticsPerPeriod(combined_df_decade)

        # Totaalstatistieken voor gebruik in de plot en het geopackage
        stats = stats_dag_per_periode["totaal"]
        stats_dec = stats_dec_per_periode["totaal"]

        # Add the cumulative discharges to the dataframe
        combined_df_cum = AddCumulative(combined_df)
        combined_df_decade_cum = AddCumulative(combined_df_decade, decade=True)

        # --- Deal with the daily values ---
        full_title = " - ".join(existing_measurements)
        fig_name = full_title.split(" - ")[0]
        meetreeks_naam = fig_name  # kolomnaam in tijdsreeks-Excel
        bron_meting = None if apply_for_water_authority is not None else waterschap
        PlotAndSave(
            combined_df=combined_df_cum,
            stats=stats,
            koppelinfo=full_title,
            fig_name=fig_name,
            bron_meting=bron_meting,
            output_folder=Path(model_folder, "results", "figures"),
            criteria_grenzen=criteria_grenzen,
        )
        fig_name_clean = _sanitize_filename(fig_name)
        pop_up_figure = f'<img src="../figures/{waterschap}/{fig_name_clean}.png" width=400 height=300>'

        # Fractie plot (optioneel — enkelvoudige link + model vereist)
        if _frac_available and isinstance(link, int):
            fractie_basin_node_id = _find_basin_for_link(model, link, _graph=_model_graph)
            if fractie_basin_node_id is not None:
                pop_up_fractie = _PlotAndSaveFractie(
                    combined_df=combined_df,
                    concentration_path=_concentration_path,
                    basin_node_id=fractie_basin_node_id,
                    tracers=_tracers,
                    fig_name=fig_name,
                    waterschap=waterschap,
                    output_folder=_frac_fig_folder,
                )
                if pop_up_fractie is not None:
                    _fractie_records.append(
                        {
                            "MeetreeksC": existing_measurements[0],
                            "Waterschap": waterschap,
                            "Aan/Af": meetlocaties_link.iloc[0]["Aan/Af"],
                            "link_id": link,
                            "basin_node_id": fractie_basin_node_id,
                            "figure_path": pop_up_fractie,
                            "QGIS_map_tip": f"figures_fracties/{waterschap}/{fig_name_clean}_fractie.png",
                            "geometry": meetlocaties_link.iloc[0]["geometry_parsed"],
                        }
                    )
            else:
                print(f"    Geen Basin bereikbaar voor link {link} — fractieplot overgeslagen")

        # Save the resulting statistics per measurement
        for key, val in [
            ("koppelinfo", fig_name_clean),
            ("link_id", link),
            ("MeetreeksC", existing_measurements[0]),
            ("Aan/Af", meetlocaties_link.iloc[0]["Aan/Af"]),
            ("NSE", stats["NSE"]),
            ("RMSE", stats["RMSE"]),
            ("MAE", stats["MAE"]),
            ("RelBias", stats["RelBias"]),
            ("KGE", stats["KGE"]),
            ("P10_reldev", stats["P10_reldev"]),
            ("P25_reldev", stats["P25_reldev"]),
            ("P75_reldev", stats["P75_reldev"]),
            ("P90_reldev", stats["P90_reldev"]),
            ("geometry", meetlocaties_link.iloc[0]["geometry_parsed"]),
            ("waterschap", waterschap),
            ("figure_path", pop_up_figure),
            ("QGIS_map_tip", f"figures/{waterschap}/{fig_name_clean}.png"),
        ]:
            results_measurements[waterschap][key].append(val)

        # --- Save the results per decade ---
        fig_name_dec = fig_name + "_decade"
        PlotAndSave(
            combined_df=combined_df_decade_cum,
            stats=stats_dec,
            koppelinfo=full_title,
            fig_name=fig_name_dec,
            bron_meting=bron_meting,
            output_folder=Path(model_folder) / "results" / "figures",
            criteria_grenzen=criteria_grenzen,
        )
        fig_name_dec_clean = _sanitize_filename(fig_name) + "_decade"
        pop_up_figure_dec = f'<img src="../figures/{waterschap}/{fig_name_dec_clean}.png" width=400 height=300>'

        for key, val in [
            ("koppelinfo", fig_name_dec_clean),
            ("link_id", link),
            ("MeetreeksC", existing_measurements[0]),
            ("Aan/Af", meetlocaties_link.iloc[0]["Aan/Af"]),
            ("NSE", stats_dec["NSE"]),
            ("RMSE", stats_dec["RMSE"]),
            ("MAE", stats_dec["MAE"]),
            ("RelBias", stats_dec["RelBias"]),
            ("KGE", stats_dec["KGE"]),
            ("P10_reldev", stats_dec["P10_reldev"]),
            ("P25_reldev", stats_dec["P25_reldev"]),
            ("P75_reldev", stats_dec["P75_reldev"]),
            ("P90_reldev", stats_dec["P90_reldev"]),
            ("geometry", meetlocaties_link.iloc[0]["geometry_parsed"]),
            ("waterschap", waterschap),
            ("figure_path", pop_up_figure_dec),
            ("QGIS_map_tip", f"figures/{waterschap}/{fig_name_dec_clean}.png"),
        ]:
            results_measurements_decade[waterschap][key].append(val)

        # --- Tijdsreeksen wegschrijven per meetlocatie ---
        ts_folder = Path(model_folder) / "results" / "tijdreeksen" / waterschap
        SaveTimeseries(
            combined_df=combined_df,
            combined_df_decade=combined_df_decade,
            fig_name_clean=fig_name_clean,
            meetreeks_naam=meetreeks_naam,
            output_folder=ts_folder,
        )

        # --- Sla record op voor het Excel-criteriabestand ---
        results_excel[waterschap].append(
            {
                "locatie": fig_name_clean,
                "stats_dag": stats_dag_per_periode,
                "stats_dec": stats_dec_per_periode,
            }
        )

    results_combined = []
    results_dec_combined = []

    # Save the results in a geopackage per waterboard
    for waterschap, results in results_measurements.items():
        results_gdf = gpd.GeoDataFrame(results, geometry="geometry")
        results_gdf.set_crs(epsg="28992", inplace=True)
        results_combined.append(results_gdf)
        results_gdf.to_file(Path(model_folder) / "results" / "Validatie_resultaten.gpkg", layer=waterschap)

    for waterschap, results in results_measurements_decade.items():
        results_gdf = gpd.GeoDataFrame(results, geometry="geometry")
        results_gdf.set_crs(epsg="28992", inplace=True)
        results_dec_combined.append(results_gdf)
        results_gdf.to_file(Path(model_folder) / "results" / "Validatie_resultaten_dec.gpkg", layer=waterschap)

    # Save all the results in one geopackage to make handling in QGIS or HTML easier
    if save_results_combined:
        final_gdf = pd.concat(results_combined, ignore_index=True)
        final_dec_gdf = pd.concat(results_dec_combined, ignore_index=True)

        final_gdf.to_file(Path(model_folder) / "results" / "Validatie_resultaten_all.gpkg", layer="Compleet")
        final_dec_gdf.to_file(Path(model_folder) / "results" / "Validatie_resultaten_dec_all.gpkg", layer="Compleet")

    # Schrijf het Excel-criteriabestand weg
    excel_path = Path(model_folder) / "results" / "Validatie_criteria.xlsx"
    ExportToExcel(results_excel, excel_path, criteria_grenzen=criteria_grenzen, beoor_kleuren=beoor_kleuren)
    print(f"Criteria Excel opgeslagen: {excel_path}")

    # ── Schrijf Fractie_locaties.gpkg met alleen locaties waarvoor een fractieplot is gemaakt ──
    if _frac_available and _fractie_records:
        frac_gdf = gpd.GeoDataFrame(_fractie_records, geometry="geometry")
        frac_gdf.set_crs(epsg=28992, inplace=True)
        frac_gpkg = Path(model_folder) / "results" / "Fractie_locaties.gpkg"
        frac_gdf.to_file(frac_gpkg, layer="fractie_locaties")
        print(f"Fractie locaties geopackage opgeslagen: {frac_gpkg} ({len(_fractie_records)} locaties)")


def ConvertToDecade(combined_df_results):
    def get_decade(ts) -> int:
        day = ts.day
        if day <= 10:
            return 1
        elif day <= 20:
            return 2
        else:
            return 3

    combined_df_results["Time"] = pd.to_datetime(combined_df_results["time"])  # ensure Time is datetime

    combined_df_results["year"] = combined_df_results["time"].dt.year
    combined_df_results["month"] = combined_df_results["time"].dt.month
    combined_df_results["decade"] = combined_df_results["time"].apply(get_decade)

    # Group by year-month-decade
    grouped = combined_df_results.groupby(["year", "month", "decade"], as_index=False).agg(
        {
            "flow_rate": "mean",  # or 'sum', depending on what you want
            "sum": "mean",  # adjust aggregation as needed
        }
    )

    # Optional: combine into a proper timestamp (e.g. midpoint of the decade)
    def build_decade_date(row) -> pd.Timestamp:
        day = {1: 1, 2: 11, 3: 21}[row["decade"]]
        return pd.Timestamp(year=int(row["year"]), month=int(row["month"]), day=day)

    grouped["time"] = grouped.apply(build_decade_date, axis=1)

    # Reorder and set dtypes to match original
    result = grouped[["time", "flow_rate", "sum"]].copy()
    result["time"] = pd.to_datetime(result["time"])
    result["flow_rate"] = result["flow_rate"].astype("float64")
    result["sum"] = result["sum"].astype("float64")

    return result


def LoadMeasurements(meas_folder) -> dict[str, pd.DataFrame]:
    """The function `LoadMeasurements` reads measurement files from a specified folder, parses date columns, and returns a dictionary of measurements.

    Parameters
    ----------
    meas_folder
        The `meas_folder` is a string that represents the folder path where the measurement files are located.

    Returns
    -------
        The function `LoadMeasurements` returns a dictionary `measurements` containing different types of
    measurements loaded from CSV files located in the specified `meas_folder`. The keys in the
    dictionary correspond to the types of measurements (e.g., 'aanvoer_dag', 'aanvoer_decade',
    'afvoer_dag', 'afvoer_decade'), and the values are pandas DataFrames.

    """
    # Define the different measurement files
    meas_files = {
        "aanvoer_dag": "Metingen_aanvoer_dag_totaal.csv",
        #'aanvoer_decade': 'Metingen_aanvoer_decade.csv',
        "afvoer_dag": "Metingen_afvoer_dag_totaal.csv",
        #'afvoer_decade':  'Metingen_afvoer_decade.csv'
    }

    measurements = {}
    for key, file in meas_files.items():
        try:
            measurements[key] = pd.read_csv(Path(meas_folder) / file, parse_dates=["Unnamed: 0"])
            measurements[key].rename(columns={"Unnamed: 0": "time"}, inplace=True)
        except:  # noqa: E722 TODO: specify exception
            try:
                measurements[key] = pd.read_csv(Path(meas_folder) / file, parse_dates=["Datum"])
                measurements[key].rename(columns={"Datum": "time"}, inplace=True)
            except ValueError:
                print("Cannot identify date/time column. Can be [Unnamed: 0, Datum] ")

    return measurements


def PlotAndSave(combined_df, stats, koppelinfo, fig_name, bron_meting, output_folder, criteria_grenzen):
    """Plots data from a pd.DataFrame, adds statistical information, and saves the plot as an image in a specified folder.

    Parameters
    ----------
    combined_df
        `combined_df` is a pd.DataFrame containing data to be plotted.
    stats
        The `stats` in the `PlotAndSave` function parameter contains statistical information (NSE, KGE, RelBias, RMSE, MAE,
        percentielen) calculated for the data being plotted.
    koppelinfo
        The `koppelinfo` parameter in the `PlotAndSave` function supplies the name of the measurement location. Used as title.
    bron_meting
        The `bron_meting` parameter in the `PlotAndSave` function represents the origin of the measurement (waterboard).
    output_folder
        The `output_folder` parameter in the `PlotAndSave` function is the directory where the plot will be
        saved. It is the location where the folder for the specific `bron_meting` will be created, and
        within that folder, the plot will be saved as a PNG file with the name supplied by 'koppelinfo'
    """
    if bron_meting is None:
        Path(output_folder).mkdir(parents=True, exist_ok=True)
    else:
        (Path(output_folder) / bron_meting).mkdir(parents=True, exist_ok=True)

    font = "Arial"

    _BEOOR_KLEUR = {
        "Goed": "#2a7a2a",
        "Matig": "#b8860b",
        "Onvoldoende": "#c0392b",
        "n.v.t.": "gray",
    }

    def _fmt(val, decimals=3):
        return f"{round(val, decimals)}" if val is not None and not np.isnan(val) else "n.v.t."

    beoordeling = BeoordeelCriteria(stats, criteria_grenzen)

    fig = plt.figure(figsize=(11, 5))
    gs = fig.add_gridspec(1, 2, width_ratios=[4, 1.1], wspace=0.20)
    ax = fig.add_subplot(gs[0])
    ax_leg = fig.add_subplot(gs[1])
    ax_leg.axis("off")

    (model_line,) = ax.plot(
        combined_df["time"], combined_df["flow_rate"], label="Ribasim", color="tab:blue", linewidth=2.0
    )
    (meas_line,) = ax.plot(combined_df["time"], combined_df["sum"], label="Meting", color="tab:orange", linewidth=2.0)
    ax.grid()
    ax.set_xticks(ticks=ax.get_xticks(), labels=ax.get_xticklabels(), rotation=45, fontname=font)
    ax.tick_params(axis="both", labelsize=11)
    for lbl in ax.get_xticklabels() + ax.get_yticklabels():
        lbl.set_fontweight("bold")
    ax.set_ylabel("Debiet [m$^3$/s]", fontdict={"fontsize": 13, "fontname": font, "fontweight": "bold"})
    ax.set_title(koppelinfo, fontname=font, wrap=True, fontsize=11)

    ax2 = ax.twinx()
    (model_cum_line,) = ax2.plot(
        combined_df["time"],
        combined_df["flow_rate_cum"] / 1_000_000,
        linestyle="--",
        color="#0047b3",
        label="Rib. cum.",
        linewidth=2.0,
    )
    (meas_cum_line,) = ax2.plot(
        combined_df["time"],
        combined_df["sum_cum"] / 1_000_000,
        linestyle="--",
        color="#cc6600",
        label="Meting cum.",
        linewidth=2.0,
    )
    ax2.set_ylabel("Cum. debiet [Mm$^3$]", fontsize=13, fontweight="bold")
    ax2.tick_params(axis="y", labelsize=11)
    for lbl in ax2.get_yticklabels():
        lbl.set_fontweight("bold")
    lines = [model_line, meas_line, model_cum_line, meas_cum_line]
    labels = [line.get_label() for line in lines]
    ax.legend(lines, labels, prop={"family": font}, fontsize=13, loc="best")

    # ── Statistieken-panel rechts ─────────────────────────────────────────────
    ax_leg.add_patch(
        plt.matplotlib.patches.FancyBboxPatch(
            (0.02, 0.02),
            0.96,
            0.96,
            transform=ax_leg.transAxes,
            boxstyle="round,pad=0.02",
            facecolor="whitesmoke",
            edgecolor="lightgray",
            linewidth=0.8,
            zorder=0,
        )
    )

    LINE_HEIGHT = 0.050  # regelafstand voor score-tekst
    VERDICT_STEP = 0.020  # beoordeling dicht onder de score
    GAP_AFTER = 0.030  # ruimte na beoordeling, vóór de volgende criteria
    SEP = 0.028  # sectie-scheiding
    y = 0.96

    def _regel(tekst, kleur="black", bold=False, suffix=None, suffix_kleur="black"):
        """Score-tekst op huidige y; beoordeling direct daarna (VERDICT_STEP).

        Gevolgd door GAP_AFTER als witruimte vóór de volgende criteria.
        """
        nonlocal y
        gewicht = "bold" if bold else "normal"
        ax_leg.text(
            0.06,
            y,
            tekst,
            transform=ax_leg.transAxes,
            fontsize=10,
            fontfamily=font,
            color=kleur,
            fontweight=gewicht,
            verticalalignment="top",
            clip_on=True,
        )
        y -= LINE_HEIGHT
        if suffix:
            ax_leg.text(
                0.12,
                y,
                suffix,
                transform=ax_leg.transAxes,
                fontsize=9.5,
                fontfamily=font,
                color=suffix_kleur,
                fontweight="bold",
                verticalalignment="top",
                clip_on=True,
            )
            y -= VERDICT_STEP
            y -= GAP_AFTER

    _regel("Statistieken (totaal)", bold=True)

    # Criteria met kleurcode
    for sleutel, label, fmt in [
        ("NSE", "NSE", ".2f"),
        ("KGE", "KGE", ".2f"),
    ]:
        val = stats.get(sleutel)
        val_str = f"{val:{fmt}}" if val is not None and not np.isnan(val) else "n.v.t."
        beoor = beoordeling.get(sleutel, "n.v.t.")
        kleur = _BEOOR_KLEUR.get(beoor, "gray")
        _regel(f"{label}: {val_str}", suffix=beoor, suffix_kleur=kleur)

    bias_val = stats.get("RelBias")
    bias_str = f"{bias_val:.1f}" if bias_val is not None and not np.isnan(bias_val) else "n.v.t."
    mean_sim_val = stats.get("mean_sim")
    mean_obs_val = stats.get("mean_obs")
    mean_sim_str = f"{mean_sim_val:.2f}" if mean_sim_val is not None and not np.isnan(mean_sim_val) else "n.v.t."
    mean_obs_str = f"{mean_obs_val:.2f}" if mean_obs_val is not None and not np.isnan(mean_obs_val) else "n.v.t."
    beoor_bias = beoordeling.get("Bias", "n.v.t.")
    kleur_bias = _BEOOR_KLEUR.get(beoor_bias, "gray")
    _regel(f"Bias: {bias_str} % ({mean_sim_str}/{mean_obs_str})", suffix=beoor_bias, suffix_kleur=kleur_bias)

    y -= SEP

    _regel("Percentielen (sim / obs)", bold=True)
    for p, sleutel in [("P10", "P10"), ("P25", "P25"), ("P75", "P75"), ("P90", "P90")]:
        sim_str = _fmt(stats.get(f"{sleutel}_sim"), 2)
        obs_str = _fmt(stats.get(f"{sleutel}_obs"), 2)
        beoor = beoordeling.get(sleutel, "n.v.t.")
        kleur = _BEOOR_KLEUR.get(beoor, "gray")
        _regel(f"{p}: {sim_str} / {obs_str}", suffix=beoor, suffix_kleur=kleur)

    fig_name_clean = _sanitize_filename(fig_name)
    if bron_meting is None:
        fig_path = Path(output_folder) / (fig_name_clean + ".png")
    else:
        fig_path = Path(output_folder) / bron_meting / (fig_name_clean + ".png")
    fig.savefig(fig_path, bbox_inches="tight", dpi=300)
    plt.close()


def ExtraInfoToevoegenAllData(
    loc_koppeltabel,
    meas_folder,
    model_folder,
    apply_for_water_authority=None,
    stat_cols=("abs_q95", "abs_q05"),
    threshold=5,
    new_col="P95_P05_alle_beschikbare_data",
    mode="any",
    above_operator=">=",
):
    """Voegt P95_P05_alle_beschikbare_data toe aan bestaande validatie-geopackages.

    Stap 1 — classificeer per (MeetreeksC, Aan/Af) op basis van de volledige meetreeks:
      - '{above_operator}{threshold}' als de conditie geldt (mode='any'/'all')
      - '<{threshold}' anders
      - 'geen_data' als de meetreeks niet in de CSV staat
    Stap 2 — voeg classificatie toe aan koppeltabel en merge met geopackages.
    """
    if isinstance(stat_cols, str):
        stat_cols = list(stat_cols)

    label_above = f"{above_operator}{threshold}"
    label_below = f"<{threshold}" if above_operator == ">=" else f"<={threshold}"

    koppeltabel = LaadKoppeltabel(loc_koppeltabel, apply_for_water_authority)
    measurements = LoadMeasurements(meas_folder)

    # --- Bereken q05 / q95 per MeetreeksC over de volledige meetreeks ---
    def _calc_stats(df):
        cols = [c for c in df.columns if c != "time"]
        q = df[cols].quantile([0.05, 0.95]).T
        q.columns = ["q05", "q95"]
        q["abs_q05"] = q["q05"].abs()
        q["abs_q95"] = q["q95"].abs()
        return q

    stats_aanvoer = _calc_stats(measurements["aanvoer_dag"])
    stats_afvoer = _calc_stats(measurements["afvoer_dag"])

    # --- Stap 1: voeg classificatie toe aan koppeltabel op (MeetreeksC, Aan/Af) ---
    def _classify(row):
        mc, aan_af = row["MeetreeksC"], row["Aan/Af"]
        stats = stats_afvoer if aan_af in ("Afvoer", "Aan&Af") else stats_aanvoer
        if mc not in stats.index:
            return "geen_data"
        available = [c for c in stat_cols if c in stats.columns]
        if not available:
            return "geen_data"
        vals = stats.loc[mc, available]
        hits = vals >= threshold if above_operator == ">=" else vals > threshold
        return label_above if (hits.any() if mode == "any" else hits.all()) else label_below

    koppeltabel[new_col] = koppeltabel.apply(_classify, axis=1)

    # --- Stap 2: merge koppeltabel (met classificatie) naar geopackages op (MeetreeksC, Aan/Af) ---
    merge_cols = koppeltabel[["MeetreeksC", "Aan/Af", new_col]]

    results_path = Path(model_folder) / "results"

    def _update_gpkg(gpkg_path: Path):
        if not gpkg_path.exists():
            print(f"Geopackage niet gevonden, overgeslagen: {gpkg_path}")
            return
        layers = gpd.list_layers(gpkg_path)["name"].tolist()
        for layer in layers:
            gdf = gpd.read_file(gpkg_path, layer=layer)
            gdf = gdf.merge(merge_cols, on=["MeetreeksC", "Aan/Af"], how="left")
            gdf[new_col] = gdf[new_col].fillna("geen_data")
            gdf.to_file(gpkg_path, layer=layer)
        print(f"Bijgewerkt: {gpkg_path.name} ({len(layers)} lagen)")

    for gpkg_name in [
        "Validatie_resultaten.gpkg",
        "Validatie_resultaten_dec.gpkg",
        "Validatie_resultaten_all.gpkg",
        "Validatie_resultaten_dec_all.gpkg",
    ]:
        _update_gpkg(results_path / gpkg_name)

    print("ExtraInfoToevoegenAllData voltooid.")


def _build_model_graph(model) -> dict:
    """Bouw eenmalig O(1)-opzoekstructuren voor link/node traversal.

    Dient éénmalig aangeroepen te worden vóór een loop over meerdere links, zodat
    elke aanroep van _find_basin_for_link O(1) dictionary-lookups gebruikt in plaats
    van O(n) DataFrame-scans.

    Returns
    -------
    dict met sleutels:
        node_type_map  - node_id  → node_type (str)
        inflow_count   - node_id  → aantal instromen (int)
        upstream_map   - to_node_id   → from_node_id
        downstream_map - from_node_id → to_node_id
        link_endpoints - link_id  → (from_node_id, to_node_id)
    """
    from collections import Counter

    links_df = model.link.df
    if links_df.index.name == "link_id":
        links_df = links_df.reset_index()

    from_ids = links_df["from_node_id"].astype(int)
    to_ids = links_df["to_node_id"].astype(int)
    link_ids = links_df["link_id"].astype(int)

    node_type_map: dict[int, str] = {}
    try:
        nodes_df = model.node.df
        if nodes_df is not None and "node_type" in nodes_df.columns:
            if nodes_df.index.name == "node_id":
                node_type_map = nodes_df["node_type"].astype(str).to_dict()
            elif "node_id" in nodes_df.columns:
                node_type_map = dict(
                    zip(nodes_df["node_id"].astype(int), nodes_df["node_type"].astype(str), strict=False)
                )
    except AttributeError:
        pass

    return {
        "node_type_map": node_type_map,
        "inflow_count": dict(Counter(to_ids)),
        "upstream_map": dict(zip(to_ids, from_ids, strict=False)),
        "downstream_map": dict(zip(from_ids, to_ids, strict=False)),
        "link_endpoints": dict(zip(link_ids, zip(from_ids, to_ids, strict=False), strict=False)),
    }


def _find_basin_for_link(model, link_id: int, max_hops: int = 20, _graph: dict | None = None) -> int | None:
    """Zoek het Basin knooppunt dat hoort bij een link_id.

    Loopt eerst stroomopwaarts (via from_node_id) totdat een Basin gevonden is of een
    Junction aangetroffen wordt. Bij een Junction of mislukking wordt stroomafwaarts
    gezocht (via to_node_id van de originele link). Een Junction stroomafwaarts
    betekent dat geen Basin gevonden kan worden → returns None.

    Parameters
    ----------
    _graph
        Pre-gebouwde opzoekstructuren van _build_model_graph. Als None wordt het
        graph éénmalig intern aangemaakt (handig voor losse aanroepen, maar gebruik
        _build_model_graph vóór een loop voor O(1) performance per aanroep).
    """
    if _graph is None:
        _graph = _build_model_graph(model)

    node_type_map = _graph["node_type_map"]
    inflow_count = _graph["inflow_count"]
    upstream_map = _graph["upstream_map"]
    downstream_map = _graph["downstream_map"]
    link_endpoints = _graph["link_endpoints"]

    if link_id not in link_endpoints:
        return None
    orig_from, orig_to = link_endpoints[link_id]

    # Stroomopwaarts: begin bij from_node_id
    node_id = orig_from
    visited: set[int] = set()
    for _ in range(max_hops):
        if node_id in visited:
            break
        visited.add(node_id)
        node_type = node_type_map.get(node_id)
        if node_type is None:
            break
        if node_type == "Basin":
            return node_id
        if node_type == "Junction" and inflow_count.get(node_id, 0) > 1:
            break
        node_id = upstream_map.get(node_id)
        if node_id is None:
            break

    # Stroomafwaarts: begin bij to_node_id van originele link
    node_id = orig_to
    visited = set()
    for _ in range(max_hops):
        if node_id in visited:
            break
        visited.add(node_id)
        node_type = node_type_map.get(node_id)
        if node_type is None:
            break
        if node_type == "Basin":
            return node_id
        if node_type == "Junction":
            return None
        node_id = downstream_map.get(node_id)
        if node_id is None:
            break

    return None


# %%
###################################################################################
###################################################################################
###################################################################################
###################################################################################
#                   LHM 4.1-vergelijking
#    Codes voor aparte analyse van Ribasim uitvoer en LHM 4.1 vergelijking
#    Hiervoor draaien we eerst de standaard nabewerking.
###################################################################################
###################################################################################
###################################################################################


def _lhm41_laad_decade_tijdreeks(ts_root: str | Path, waterschap: str, fig_name_clean: str) -> pd.DataFrame | None:
    path = Path(ts_root) / waterschap / f"{fig_name_clean}.xlsx"
    if not path.exists():
        print(f"  Tijdreeks niet gevonden: {path}")
        return None
    df = pd.read_excel(path, sheet_name="Decade")
    df = df.iloc[:, :3].copy()
    df.columns = ["Datum", "meting", "Ribasim"]
    df["Datum"] = pd.to_datetime(df["Datum"])
    df["meting"] = pd.to_numeric(df["meting"], errors="coerce")
    df["Ribasim"] = pd.to_numeric(df["Ribasim"], errors="coerce")
    return df


def _lhm41_laad_csv(lhm41_folder: str | Path, csv_naam: str, LHM41_TEKEN_CORRECTIE: dict | None) -> pd.DataFrame | None:
    path = Path(lhm41_folder) / csv_naam
    if not path.exists():
        print(f"  LHM 4.1 CSV niet gevonden: {path}")
        return None
    df = pd.read_csv(path, sep=None, engine="python")
    df = df.iloc[:, :2].copy()
    df.columns = ["Datum", "LHM41"]
    df["Datum"] = pd.to_datetime(df["Datum"])
    df["LHM41"] = pd.to_numeric(df["LHM41"], errors="coerce")
    if LHM41_TEKEN_CORRECTIE is not None:
        factor = LHM41_TEKEN_CORRECTIE.get(csv_naam)
        if factor is not None:
            df["LHM41"] *= factor
            print(f"  Tekencorrectie toegepast op {csv_naam} (factor: {factor})")
    return df


def _lhm41_bereken_p95_klasse(
    measurements: dict,
    meetreeks_c: str,
    aan_af: str,
    threshold: float = 5.0,
) -> str:
    meas_df = measurements["afvoer_dag"] if aan_af in ("Afvoer", "Aan&Af") else measurements["aanvoer_dag"]
    if meetreeks_c not in meas_df.columns:
        return "geen_data"
    serie = meas_df[meetreeks_c].dropna()
    if len(serie) == 0:
        return "geen_data"
    q05 = serie.quantile(0.05)
    q95 = serie.quantile(0.95)
    return ">=5" if (abs(q05) >= threshold or abs(q95) >= threshold) else "<5"


def _lhm41_bereken_p95_klasse_som(
    measurements: dict,
    locs: list,
    threshold: float = 5.0,
) -> str:
    aan_af = locs[0]["Aan/Af"]
    meas_df = measurements["afvoer_dag"] if aan_af in ("Afvoer", "Aan&Af") else measurements["aanvoer_dag"]

    series_list = []
    for loc in locs:
        col = loc["MeetreeksC"]
        if col in meas_df.columns:
            series_list.append(meas_df[col].rename(col))

    if not series_list:
        return "geen_data"

    combined = pd.concat(series_list, axis=1).dropna()
    if combined.empty:
        return "geen_data"

    som = combined.sum(axis=1)
    q05 = som.quantile(0.05)
    q95 = som.quantile(0.95)
    return ">=5" if (abs(q05) >= threshold or abs(q95) >= threshold) else "<5"


def _lhm41_cum_afwijking_jaarlijks(
    df: pd.DataFrame,
    col_model: str,
    col_obs: str,
    zomer: bool = False,
) -> dict:
    leeg = {"per_jaar": {}, "gemiddeld": np.nan, "totaal": np.nan}

    d = df.copy()
    if zomer:
        d = d[d["Datum"].dt.month.between(4, 9)]
    if d.empty:
        return leeg

    d["jaar"] = d["Datum"].dt.year
    afwijkingen_per_jaar = {}

    for jaar, grp in d.groupby("jaar"):
        grp_geldig = grp.dropna(subset=[col_obs])
        if grp_geldig.empty:
            continue
        laatste_datum = grp_geldig["Datum"].max()
        grp_tot_einde = grp[grp["Datum"] <= laatste_datum]

        cum_obs = grp_tot_einde[col_obs].fillna(0).sum()
        cum_mod = grp_tot_einde[col_model].fillna(0).sum()

        if abs(cum_obs) > 0:
            afwijkingen_per_jaar[int(str(jaar))] = round(abs(1 - cum_mod / cum_obs) * 100, 1)

    waarden = list(afwijkingen_per_jaar.values())
    gemiddeld = float(np.nanmean(waarden)) if waarden else np.nan

    d_geldig = d.dropna(subset=[col_obs])
    if not d_geldig.empty:
        laatste_datum_totaal = d_geldig["Datum"].max()
        d_tot_einde = d[d["Datum"] <= laatste_datum_totaal]
        cum_obs_tot = d_tot_einde[col_obs].fillna(0).sum()
        cum_mod_tot = d_tot_einde[col_model].fillna(0).sum()
        totaal = round(abs(1 - cum_mod_tot / cum_obs_tot) * 100, 1) if abs(cum_obs_tot) > 0 else np.nan
    else:
        totaal = np.nan

    return {"per_jaar": afwijkingen_per_jaar, "gemiddeld": gemiddeld, "totaal": totaal}


def _lhm41_bereken_statistieken(
    df: pd.DataFrame,
    col_model: str,
    col_obs: str,
) -> dict:
    df_stats = df.rename(columns={col_model: "flow_rate", col_obs: "sum", "Datum": "time"})

    stats_per_periode = GetStatisticsPerPeriod(df_stats)
    stats = dict(stats_per_periode["totaal"])

    nse_per_jaar = {
        jaar: s["NSE"]
        for jaar, s in stats_per_periode.items()
        if jaar != "totaal" and s is not None and not np.isnan(s.get("NSE", np.nan))
    }
    stats["NSE_per_jaar"] = nse_per_jaar

    cum_jaar = _lhm41_cum_afwijking_jaarlijks(df, col_model, col_obs, zomer=False)
    cum_zomer = _lhm41_cum_afwijking_jaarlijks(df, col_model, col_obs, zomer=True)

    stats["CumJaar"] = cum_jaar["gemiddeld"]
    stats["CumJaar_totaal"] = cum_jaar["totaal"]
    stats["CumJaar_per_jaar"] = cum_jaar["per_jaar"]
    stats["CumZomer"] = cum_zomer["gemiddeld"]
    stats["CumZomer_totaal"] = cum_zomer["totaal"]
    stats["CumZomer_per_jaar"] = cum_zomer["per_jaar"]

    return stats


def _lhm41_toets_locaties(
    stats_lijst: list[dict],
    categorie: str,
    p95_klasse: str,
    criteria_lhm41: dict,
) -> list[dict]:
    criteria = criteria_lhm41.get((categorie, p95_klasse), [])
    rijen = []
    for type_stat, drempel, pct_vereist, label in criteria:
        waarden = [s[type_stat] for s in stats_lijst if not np.isnan(s.get(type_stat, np.nan))]
        n_totaal = len(waarden)
        if n_totaal == 0:
            rijen.append(
                {
                    "Criterium": label,
                    "N": 0,
                    "N_voldoet": 0,
                    "Pct_voldoet": np.nan,
                    "Pct_vereist": pct_vereist,
                    "Beoordeling": "n.v.t.",
                }
            )
            continue
        n_voldoet = sum(v > drempel for v in waarden) if type_stat == "NSE" else sum(v < drempel for v in waarden)
        pct = n_voldoet / n_totaal * 100
        rijen.append(
            {
                "Criterium": label,
                "N": n_totaal,
                "N_voldoet": n_voldoet,
                "Pct_voldoet": round(pct, 1),
                "Pct_vereist": pct_vereist,
                "Beoordeling": "Goed" if pct >= pct_vereist else "Onvoldoende",
            }
        )
    return rijen


def _lhm41_stats_regels(
    stats_rib: dict, stats_lhm: dict, categorie: str, p95_klasse: str, criteria_lhm41: dict
) -> list[dict]:
    criteria = criteria_lhm41.get((categorie, p95_klasse), [])

    def _beoordeling(waarde: float, type_stat: str, drempel: float) -> tuple[str, str]:
        if np.isnan(waarde):
            return "n.b.", "gray"
        voldoet = waarde > drempel if type_stat == "NSE" else waarde < drempel
        return ("Voldoet", "#2a7a2a") if voldoet else ("Voldoet niet", "#c0392b")

    regels = []
    regels.append(
        {"text": f"Categorie: {categorie}  |  P95: {p95_klasse} m³/s", "color": None, "bold": True, "indent": False}
    )
    regels.append({"text": "", "color": None, "bold": False, "indent": False})

    for type_stat, drempel, pct_vereist, label in criteria:
        rib_val = stats_rib.get(type_stat, np.nan)
        lhm_val = stats_lhm.get(type_stat, np.nan)
        rib_str = f"{rib_val:.1f}" if not np.isnan(rib_val) else "n.b."
        lhm_str = f"{lhm_val:.1f}" if not np.isnan(lhm_val) else "n.b."
        rib_label, rib_kleur = _beoordeling(rib_val, type_stat, drempel)
        lhm_label, lhm_kleur = _beoordeling(lhm_val, type_stat, drempel)

        regels.append({"text": f"{label}  (norm ≥{pct_vereist}% loc.)", "color": None, "bold": False, "indent": False})
        regels.append(
            {
                "text": f"  Ribasim:  {rib_str}",
                "color": None,
                "bold": False,
                "indent": True,
                "suffix": rib_label,
                "suffix_color": rib_kleur,
            }
        )
        regels.append(
            {
                "text": f"  LHM 4.1:  {lhm_str}",
                "color": None,
                "bold": False,
                "indent": True,
                "suffix": lhm_label,
                "suffix_color": lhm_kleur,
            }
        )
        regels.append({"text": "", "color": None, "bold": False, "indent": False})

    for sleutel, label in [("NSE", "NSE totaal")]:
        if not any(c[0] == sleutel for c in criteria):
            rib_val = stats_rib.get(sleutel, np.nan)
            lhm_val = stats_lhm.get(sleutel, np.nan)
            rib_str = f"{rib_val:.2f}" if not np.isnan(rib_val) else "n.b."
            lhm_str = f"{lhm_val:.2f}" if not np.isnan(lhm_val) else "n.b."
            regels.append({"text": label, "color": None, "bold": False, "indent": False})
            regels.append(
                {
                    "text": f"  Ribasim: {rib_str}   LHM 4.1: {lhm_str}",
                    "color": "dimgray",
                    "bold": False,
                    "indent": True,
                }
            )
            regels.append({"text": "", "color": None, "bold": False, "indent": False})

    return regels


def _lhm41_plot_vergelijking(
    df_merged: pd.DataFrame,
    titel: str,
    output_folder: str | Path,
    fig_naam: str,
    stats_rib: dict | None = None,
    stats_lhm: dict | None = None,
    categorie: str = "",
    p95_klasse: str = "",
    criteria_lhm41: dict | None = None,
) -> None:
    if criteria_lhm41 is None:
        criteria_lhm41 = {}
    Path(output_folder).mkdir(parents=True, exist_ok=True)

    MULTIPLIER = 10 * 86400
    font = "Arial"

    datum = df_merged["Datum"]
    meting = df_merged["meting_som"]
    ribasim = df_merged["Ribasim_som"]
    lhm41 = df_merged["LHM41"]

    cum_meting = meting.fillna(0).cumsum() * MULTIPLIER / 1_000_000
    cum_ribasim = ribasim.fillna(0).cumsum() * MULTIPLIER / 1_000_000
    cum_lhm41 = lhm41.fillna(0).cumsum() * MULTIPLIER / 1_000_000

    heeft_stats = stats_rib is not None and stats_lhm is not None
    fig_breedte = 12 if heeft_stats else 8

    fig = plt.figure(figsize=(fig_breedte, 7))
    if heeft_stats:
        gs = fig.add_gridspec(2, 2, width_ratios=[3, 1.8], hspace=0.15, wspace=0.12)
        ax_dec = fig.add_subplot(gs[0, 0])
        ax_cum = fig.add_subplot(gs[1, 0], sharex=ax_dec)
        ax_leg = fig.add_subplot(gs[:, 1])
        ax_leg.axis("off")
    else:
        gs = fig.add_gridspec(2, 1, hspace=0.15)
        ax_dec = fig.add_subplot(gs[0])
        ax_cum = fig.add_subplot(gs[1], sharex=ax_dec)

    fig.suptitle(titel, fontname=font, fontsize=12, fontweight="bold", wrap=True)

    ax_dec.plot(datum, meting, label="Meting", color="tab:orange", linewidth=2.0)
    ax_dec.plot(datum, ribasim, label="Ribasim", color="tab:blue", linewidth=2.0)
    ax_dec.plot(datum, lhm41, label="LHM 4.1", color="tab:green", linewidth=2.0, linestyle="--")
    ax_dec.set_ylabel("Debiet [m³/s]", fontdict={"fontsize": 12, "fontname": font, "fontweight": "bold"})
    ax_dec.legend(fontsize=10, loc="upper right")
    ax_dec.grid(True, linewidth=0.5)
    ax_dec.tick_params(axis="x", labelbottom=False)
    ax_dec.tick_params(axis="y", labelsize=11)
    for lbl in ax_dec.get_yticklabels():
        lbl.set_fontweight("bold")

    ax_cum.plot(datum, cum_meting, label="Meting cum.", color="tab:orange", linewidth=2.0)
    ax_cum.plot(datum, cum_ribasim, label="Ribasim cum.", color="tab:blue", linewidth=2.0)
    ax_cum.plot(datum, cum_lhm41, label="LHM 4.1 cum.", color="tab:green", linewidth=2.0, linestyle="--")
    ax_cum.set_ylabel("Cum. debiet [Mm³]", fontdict={"fontsize": 12, "fontname": font, "fontweight": "bold"})
    ax_cum.legend(fontsize=10, loc="upper left")
    ax_cum.grid(True, linewidth=0.5)
    ax_cum.tick_params(axis="x", rotation=45, labelsize=11)
    ax_cum.tick_params(axis="y", labelsize=11)
    for lbl in ax_cum.get_xticklabels() + ax_cum.get_yticklabels():
        lbl.set_fontweight("bold")

    if heeft_stats:
        regels = _lhm41_stats_regels(stats_rib, stats_lhm, categorie, p95_klasse, criteria_lhm41)
        ax_leg.add_patch(
            plt.matplotlib.patches.FancyBboxPatch(
                (0.02, 0.02),
                0.96,
                0.96,
                transform=ax_leg.transAxes,
                boxstyle="round,pad=0.02",
                facecolor="whitesmoke",
                edgecolor="lightgray",
                linewidth=0.8,
                zorder=0,
            )
        )
        LINE_HEIGHT = 0.050
        VERDICT_STEP = 0.020
        GAP_AFTER = 0.030
        y = 0.96

        for regel in regels:
            tekst = regel["text"]
            kleur = regel["color"] or "black"
            suffix = regel.get("suffix")
            suffix_kleur = regel.get("suffix_color", "black")
            ax_leg.text(
                0.06,
                y,
                tekst,
                transform=ax_leg.transAxes,
                fontsize=10.5,
                fontfamily=font,
                color=kleur,
                fontweight="bold",
                verticalalignment="top",
                clip_on=True,
            )
            y -= LINE_HEIGHT
            if suffix:
                ax_leg.text(
                    0.08,
                    y,
                    suffix,
                    transform=ax_leg.transAxes,
                    fontsize=10.5,
                    fontfamily=font,
                    color=suffix_kleur,
                    fontweight="bold",
                    verticalalignment="top",
                    clip_on=True,
                )
                y -= VERDICT_STEP
                y -= GAP_AFTER

    fig.savefig(Path(output_folder) / f"{fig_naam}.png", dpi=300, bbox_inches="tight")
    plt.close(fig)


def _lhm41_kleur_toetsing(ws, df: pd.DataFrame) -> None:
    if "Beoordeling" not in df.columns:
        return
    col_idx = df.columns.tolist().index("Beoordeling") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = _BEOOR_KLEUREN.get(cell.value, _BEOOR_KLEUREN["n.v.t."])


def AnalyseLHM41Vergelijking(
    gpkg_koppellaag: str | Path,
    layer_naam: str,
    model_folder: str | Path,
    meas_folder: str | Path,
    lhm41_folder: str | Path,
    criteria_lhm41: dict | None = None,
    output_path: str | Path | None = None,
    LHM41_TEKEN_CORRECTIE: dict | None = None,
) -> None:
    """Voert de LHM 4.1-vergelijkingsanalyse uit en schrijft een criteria-Excel weg.

    Parameters
    ----------
    gpkg_koppellaag
        Pad naar de geopackage met de koppellaag.
    layer_naam
        Naam van de laag in de geopackage (bijv. ``'koppeling_lhm4_1_reeksen'``).
    model_folder
        Map van het Ribasim-model; tijdreeksen worden gelezen uit
        ``{model_folder}/results/tijdreeksen/``.
    meas_folder
        Map met de meetbestanden (Metingen_aanvoer/afvoer_dag_totaal.csv).
    lhm41_folder
        Map met de LHM 4.1 decade CSV-bestanden.
    criteria_lhm41
        Toetsingscriteria per (categorie, P95-klasse). Standaard: _CRITERIA_LHM41_DEFAULT.
    output_path
        Pad voor het output-Excel. Standaard:
        ``{model_folder}/results/Validatie_criteria_lhm4_1.xlsx``.
    LHM41_TEKEN_CORRECTIE
        Tekencorrectie per CSV-bestandsnaam. Standaard: geen correcties.
    """
    _CRITERIA_LHM41_DEFAULT = {
        ("Afvoer", ">=5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumJaar", 25, 80, "Cum. jaarafvoer dif. < 25%"),
        ],
        ("Afvoer", "<5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumJaar", 33, 80, "Cum. jaarafvoer dif. < 33%"),
        ],
        ("Aanvoer", ">=5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumZomer", 25, 80, "Cum. zomer aanvoer dif. < 25%"),
        ],
        ("Aanvoer", "<5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumZomer", 50, 80, "Cum. zomer aanvoer dif. < 50%"),
        ],
    }

    if criteria_lhm41 is None:
        criteria_lhm41 = _CRITERIA_LHM41_DEFAULT
    if output_path is None:
        output_path = Path(model_folder) / "results" / "Validatie_criteria_lhm4_1.xlsx"

    ts_root = Path(model_folder) / "results" / "tijdreeksen"

    # ── 1. Laad koppellaag ────────────────────────────────────────────────────
    print("Laden koppellaag...")
    gdf = gpd.read_file(gpkg_koppellaag, layer=layer_naam)
    ontbrekend = {"MeetreeksC", "Aan/Af", "Waterschap", "LHM_4_1_csv"} - set(gdf.columns)
    if ontbrekend:
        raise ValueError(f"Kolommen ontbreken in koppellaag: {ontbrekend}")
    print(f"  {len(gdf)} rijen geladen.")

    # ── 2. Laad metingen voor P95-classificatie ───────────────────────────────
    print("Laden metingen...")
    measurements = LoadMeasurements(meas_folder)

    # ── 3. Per locatie: decade-tijdreeks + statistieken Ribasim vs. meting ───
    print("Berekenen statistieken per locatie (Ribasim vs. meting)...")
    locatie_resultaten = []

    # Mapping
    _LHM41_AANAF_NAAR_CATEGORIE = {
        "Afvoer": "Afvoer",
        "Aanvoer": "Aanvoer",
        "Aan&Af": "Afvoer",
    }

    for _, rij in gdf.iterrows():
        meetreeks_c = rij["MeetreeksC"]
        aan_af = rij["Aan/Af"]
        waterschap = rij["Waterschap"]
        lhm41_csv = rij["LHM_4_1_csv"]
        categorie = _LHM41_AANAF_NAAR_CATEGORIE.get(aan_af, "Afvoer")
        p95_klasse = _lhm41_bereken_p95_klasse(measurements, meetreeks_c, aan_af)
        fig_name = _sanitize_filename(meetreeks_c)

        df_dec = _lhm41_laad_decade_tijdreeks(ts_root, waterschap, fig_name)
        if df_dec is None or df_dec.dropna(subset=["meting", "Ribasim"]).empty:
            print(f"  Geen bruikbare decade-data voor {meetreeks_c}, overgeslagen.")
            continue

        stats_rib = _lhm41_bereken_statistieken(df_dec, "Ribasim", "meting")

        locatie_resultaten.append(
            {
                "MeetreeksC": meetreeks_c,
                "Waterschap": waterschap,
                "Aan/Af": aan_af,
                "Categorie": categorie,
                "P95_klasse": p95_klasse,
                "LHM_4_1_csv": lhm41_csv,
                "df_dec": df_dec,
                "stats_rib": stats_rib,
            }
        )

    print(f"  {len(locatie_resultaten)} locaties verwerkt.")

    # ── 4. Groepeer op LHM_4_1_csv: sommeer + vergelijk met LHM 4.1 ─────────
    print("Berekenen statistieken per groep (LHM 4.1 vs. meting)...")
    groep_resultaten: list[dict[str, Any]] = []

    groepen: dict[str, list] = defaultdict(list)
    for loc in locatie_resultaten:
        csv_naam = str(loc["LHM_4_1_csv"]).strip()
        if csv_naam and csv_naam.lower() != "nan":
            groepen[csv_naam].append(loc)

    for csv_naam, locs in groepen.items():
        datum_index = locs[0]["df_dec"]["Datum"].values
        rib_som = np.zeros(len(datum_index))
        met_som = np.zeros(len(datum_index))
        has_meting = np.zeros(len(datum_index), dtype=bool)

        for loc in locs:
            df_i = loc["df_dec"].set_index("Datum").reindex(datum_index)
            rib_som += df_i["Ribasim"].fillna(0).values
            met_som += df_i["meting"].fillna(0).values
            has_meting |= df_i["meting"].notna().values

        df_som = pd.DataFrame(
            {
                "Datum": datum_index,
                "Ribasim_som": rib_som,
                "meting_som": np.where(has_meting, met_som, np.nan),
            }
        )

        df_lhm = _lhm41_laad_csv(lhm41_folder, csv_naam, LHM41_TEKEN_CORRECTIE)
        if df_lhm is None:
            continue
        df_merged = df_som.merge(df_lhm, on="Datum", how="left")

        categorie = locs[0]["Categorie"]
        p95_klasse = _lhm41_bereken_p95_klasse_som(measurements, locs)
        if len(locs) > 1:
            individueel = [loc_["P95_klasse"] for loc_ in locs]
            if any(k != p95_klasse for k in individueel):
                namen = [loc_["MeetreeksC"] for loc_ in locs]
                print(
                    f"  P95-klasse gewijzigd na optelling voor {csv_naam}: "
                    f"individueel {individueel} → gecombineerd '{p95_klasse}' "
                    f"({', '.join(namen)})"
                )
        stats_lhm = _lhm41_bereken_statistieken(df_merged, "LHM41", "meting_som")
        stats_rib_groep = _lhm41_bereken_statistieken(df_merged, "Ribasim_som", "meting_som")

        fig_naam = _sanitize_filename(Path(csv_naam).stem)
        titel = ", ".join(loc_["MeetreeksC"] for loc_ in locs)
        fig_folder = Path(model_folder) / "results" / "figures_lhm4_1_vergelijking"
        _lhm41_plot_vergelijking(
            df_merged,
            titel,
            fig_folder,
            fig_naam,
            stats_rib=stats_rib_groep,
            stats_lhm=stats_lhm,
            categorie=categorie,
            p95_klasse=p95_klasse,
            criteria_lhm41=criteria_lhm41,
        )

        groep_resultaten.append(
            {
                "LHM_4_1_csv": csv_naam,
                "MeetreeksC": ", ".join(loc_["MeetreeksC"] for loc_ in locs),
                "Waterschap": locs[0]["Waterschap"],
                "Categorie": categorie,
                "P95_klasse": p95_klasse,
                "fig_naam": fig_naam,
                "locs": locs,
                "stats_rib": stats_rib_groep,
                "stats_lhm": stats_lhm,
            }
        )

    print(f"  {len(groep_resultaten)} groepen verwerkt.")

    # ── 5. Toetsing per (Categorie, P95_klasse) ───────────────────────────────
    print("Evalueren toetsingscriteria...")

    def _maak_toetsing_df(resultaten: list, stats_sleutel: str, criteria_lhm41: dict) -> pd.DataFrame:
        per_cat: dict[tuple, list] = defaultdict(list)
        for res in resultaten:
            per_cat[(res["Categorie"], res["P95_klasse"])].append(res[stats_sleutel])
        rijen = []
        for (cat, klasse), stats_lijst in sorted(per_cat.items()):
            rijen.extend(
                {"Categorie": cat, "P95_klasse": klasse, **tr}
                for tr in _lhm41_toets_locaties(stats_lijst, cat, klasse, criteria_lhm41)
            )
        return pd.DataFrame(rijen) if rijen else pd.DataFrame()

    df_toets_rib = _maak_toetsing_df(groep_resultaten, "stats_rib", criteria_lhm41)
    df_toets_lhm = _maak_toetsing_df(groep_resultaten, "stats_lhm", criteria_lhm41)

    # ── 6. Statistieken-overzicht ─────────────────────────────────────────────
    def _stats_rij(s: dict, prefix: str) -> dict:
        rij = {
            f"{prefix}_NSE": round(s.get("NSE", np.nan), 3),
            f"{prefix}_RelBias": round(s.get("RelBias", np.nan), 1),
            f"{prefix}_CumJaar_gem": round(s.get("CumJaar", np.nan), 1),
            f"{prefix}_CumJaar_totaal": round(s.get("CumJaar_totaal", np.nan), 1),
            f"{prefix}_CumZomer_gem": round(s.get("CumZomer", np.nan), 1),
            f"{prefix}_CumZomer_totaal": round(s.get("CumZomer_totaal", np.nan), 1),
        }
        for jaar, nse in sorted(s.get("NSE_per_jaar", {}).items()):
            rij[f"{prefix}_NSE_{jaar}"] = round(nse, 3)
        for jaar, afwijk in sorted(s.get("CumJaar_per_jaar", {}).items()):
            rij[f"{prefix}_CumJaar_{jaar}"] = round(afwijk, 1)
        for jaar, afwijk in sorted(s.get("CumZomer_per_jaar", {}).items()):
            rij[f"{prefix}_CumZomer_{jaar}"] = round(afwijk, 1)
        return rij

    df_locaties = pd.DataFrame(
        [
            {
                "LHM_4_1_csv": g["LHM_4_1_csv"],
                "MeetreeksC": g["MeetreeksC"],
                "Waterschap": g["Waterschap"],
                "Categorie": g["Categorie"],
                "P95_klasse": g["P95_klasse"],
                **_stats_rij(g["stats_rib"], "Rib"),
            }
            for g in groep_resultaten
        ]
    )

    df_groepen = pd.DataFrame(
        [
            {
                "LHM_4_1_csv": g["LHM_4_1_csv"],
                "MeetreeksC": g["MeetreeksC"],
                "Categorie": g["Categorie"],
                "P95_klasse": g["P95_klasse"],
                **_stats_rij(g["stats_lhm"], "LHM41"),
            }
            for g in groep_resultaten
        ]
    )

    # ── 7. Schrijf Excel ──────────────────────────────────────────────────────
    print(f"Schrijven Excel: {output_path}")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not df_locaties.empty:
            df_locaties.to_excel(writer, sheet_name="Statistieken_Ribasim", index=False)
        if not df_groepen.empty:
            df_groepen.to_excel(writer, sheet_name="Statistieken_LHM41", index=False)
        if not df_toets_rib.empty:
            df_toets_rib.to_excel(writer, sheet_name="Toetsing_Ribasim", index=False)
            _lhm41_kleur_toetsing(writer.sheets["Toetsing_Ribasim"], df_toets_rib)
        if not df_toets_lhm.empty:
            df_toets_lhm.to_excel(writer, sheet_name="Toetsing_LHM41", index=False)
            _lhm41_kleur_toetsing(writer.sheets["Toetsing_LHM41"], df_toets_lhm)

    # ── 8. GeoPackage wegschrijven voor HTML-viewer ───────────────────────────
    print("GeoPackage wegschrijven voor HTML-viewer...")
    geom_lookup = {row["MeetreeksC"]: row.geometry for _, row in gdf.iterrows() if row.geometry is not None}

    gpkg_records = []
    for res in groep_resultaten:
        geom = None
        for loc in res["locs"]:
            geom = geom_lookup.get(loc["MeetreeksC"])
            if geom is not None:
                break
        if geom is None:
            continue

        fig_naam = res["fig_naam"]
        s_rib = res["stats_rib"]
        s_lhm = res["stats_lhm"]

        def _r(val, dec=2):
            return round(float(val), dec) if val is not None and not np.isnan(float(val)) else None

        gpkg_records.append(
            {
                "koppelinfo": res["MeetreeksC"],
                "MeetreeksC": res["MeetreeksC"],
                "Waterschap": res["Waterschap"],
                "Categorie": res["Categorie"],
                "P95_klasse": res["P95_klasse"],
                "LHM_4_1_csv": res["LHM_4_1_csv"],
                "NSE_rib": _r(s_rib.get("NSE")),
                "CumJaar_rib": _r(s_rib.get("CumJaar"), 1),
                "CumZomer_rib": _r(s_rib.get("CumZomer"), 1),
                "RelBias_rib": _r(s_rib.get("RelBias"), 1),
                "NSE_lhm": _r(s_lhm.get("NSE")),
                "CumJaar_lhm": _r(s_lhm.get("CumJaar"), 1),
                "CumZomer_lhm": _r(s_lhm.get("CumZomer"), 1),
                "RelBias_lhm": _r(s_lhm.get("RelBias"), 1),
                "figure_path": f'<img src="../figures_lhm4_1_vergelijking/{fig_naam}.png" width=400 height=300>',
                "QGIS_map_tip": f"figures_lhm4_1_vergelijking/{fig_naam}.png",
                "geometry": geom,
            }
        )

    if gpkg_records:
        gdf_out = gpd.GeoDataFrame(gpkg_records, geometry="geometry", crs=gdf.crs)
        gpkg_out = Path(model_folder) / "results" / "Validatie_resultaten_lhm41.gpkg"
        gdf_out.to_file(gpkg_out, driver="GPKG", layer="lhm41_vergelijking")
        print(f"  GeoPackage opgeslagen: {gpkg_out} ({len(gdf_out)} groepen)")
    else:
        print("  Geen groepen met geometrie gevonden, GeoPackage niet weggeschreven.")

    print("Klaar.")


# %%

if __name__ == "__main__":
    ########################################################################################################################################
    # Implementatie lokaal voor testen
    cloud = CloudStorage()
    base_koppeltabel = cloud.joinpath("Basisgegevens/resultaatvergelijking/koppeltabel_2026")

    loc_koppeltabel = (
        base_koppeltabel / "Transformed_koppeltabel_versie_RijnenIJssel_2026_5_1_Feedback_Verwerkt_HydroLogic.xlsx"
    )
    loc_specifieke_bewerking = base_koppeltabel / "Specifiek_bewerking_versieRijnenIJssel_2026_5_1.xlsx"
    waterboard = "RijnenIJssel"
    waterboard_model_versions = cloud.uploaded_models(authority=waterboard)

    latest_model_version = sorted(
        [i for i in waterboard_model_versions if i.model == waterboard], key=lambda x: getattr(x, "sorter", "")
    )[-1]

    model_folder = cloud.joinpath(f"{waterboard}/modellen", latest_model_version.path_string)
    # toml_naam = "lhm-coupled.toml"
    toml_naam = "wrij.toml"

    meas_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/meetreeksen_2026")

    # synchronize paths
    cloud.synchronize([loc_koppeltabel, loc_specifieke_bewerking, meas_folder, model_folder])

    # validatie_lhm4_1 = os.path.join(base, "Toetsing LHM 4.1", )
    ########################################################################################################################################

    # ── Toetsingscriteria en drempelwaarden ──────────────────────────────────
    # Richtwaarden: Goed = enige klasse (geen Matig); anders Onvoldoende.
    # NSE > 0.4, KGE > 0.5, |Bias| < 15%, P10/P90 binnen ±30%, P25/P75 binnen ±15%.
    CRITERIA_GRENZEN = {
        "Bias": {"Goed": 15.0},
        "NSE": {"Goed": 0.4},
        "KGE": {"Goed": 0.5},
        "P10": {"Goed": 30.0},
        "P25": {"Goed": 15.0},
        "P75": {"Goed": 15.0},
        "P90": {"Goed": 30.0},
    }
    BEOOR_KLEUREN = {
        "Goed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Matig": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Onvoldoende": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "n.v.t.": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }
    # 6 kwantitatieve criteria; model voldoet als ≥ MIN_GROEPEN_VOLDOEN ervan "Goed" zijn.
    # NSE én KGE tellen samen als één criterium (tenminste één moet voldoen).
    CRITERIA_GROEPEN = {
        "Bias": {"kolommen": ["Beoor_Bias_dec"], "min_goed": 1},
        "NSE_KGE": {"kolommen": ["Beoor_NSE_dec", "Beoor_KGE_dec"], "min_goed": 1},
        "P10": {"kolommen": ["Beoor_P10_dec"], "min_goed": 1},
        "P25": {"kolommen": ["Beoor_P25_dec"], "min_goed": 1},
        "P75": {"kolommen": ["Beoor_P75_dec"], "min_goed": 1},
        "P90": {"kolommen": ["Beoor_P90_dec"], "min_goed": 1},
    }
    MIN_GROEPEN_VOLDOEN = 3  # ≥ 3 van de 6 criteria Goed
    DREMPEL_HWS = 75.0  # HWS: 75% van locaties moet voldoen
    DREMPEL_REGIONAAL = 50.0  # Regionaal: 50% van locaties moet voldoen
    HWS_WATERSCHAP = "Rijkswaterstaat"

    # ── LHM 4.1-toetsingscriteria ────────────────────────────────────────────
    CRITERIA_LHM41 = {
        ("Afvoer", ">=5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumJaar", 25, 80, "Cum. jaarafvoer dif. < 25%"),
        ],
        ("Afvoer", "<5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumJaar", 33, 80, "Cum. jaarafvoer dif. < 33%"),
        ],
        ("Aanvoer", ">=5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumZomer", 25, 80, "Cum. zomer aanvoer dif. < 25%"),
        ],
        ("Aanvoer", "<5"): [
            ("NSE", 0.2, 80, "NSE > 0.2"),
            ("NSE", 0.5, 60, "NSE > 0.5"),
            ("CumZomer", 50, 80, "Cum. zomer aanvoer dif. < 50%"),
        ],
    }
    # Tekencorrectie per LHM 4.1 CSV-bestand (-1 als tekenconventie verschilt van Ribasim/meting)
    LHM41_TEKEN_CORRECTIE: dict[str, float] = {
        "Aanvoer Gemaal Winsemius_LHM_reeks.csv": -1.0,
    }

    RUN_COMPARE = True
    RUN_EXTRA_INFO = True
    RUN_EINDBEOORDELING = True
    RUN_LHM41 = True

    RUN_HTML_VIEWER = True
    # Dit gaat nog niet ivm hoeveelheid links, polygons en nodes in modellen
    model_gpkg_in_html = None
    # Als we ook resultaten vanuit LHM4.1 kunnen we deze laag wel of niet tonen in html viewer
    #!!!TODO: Draai hiervoor eerst notebooks/observations/analyse_lhm41_vergelijking
    include_lhm41_html = RUN_LHM41
    # Als we ook fracties kunnen plotten en tonen in html viewer
    include_fractie_html = True

    RUN_UPLOAD = False

    # ── Verwerk meetreeksen, bereken statistieken, schrijf figuren/geopackages/Excel ──
    if RUN_COMPARE:
        CompareOutputMeasurements(
            loc_koppeltabel=loc_koppeltabel,
            loc_specifics=loc_specifieke_bewerking,
            meas_folder=meas_folder,
            model_folder=model_folder,
            toml_file=toml_naam,
            filetype="flow",
            criteria_grenzen=CRITERIA_GRENZEN,
            beoor_kleuren=BEOOR_KLEUREN,
            save_results_combined=True,
            output_is_feather=False,
            output_is_nc=True,
            resample_to_daily=True,
        )
    else:
        print(
            "LET OP: RUN_COMPARE is uitgeschakeld. De overige stappen (RUN_EXTRA_INFO, "
            "RUN_EINDBEOORDELING, RUN_LHM41, RUN_HTML_VIEWER) vereisen dat "
            "CompareOutputMeasurements al eerder succesvol is gedraaid."
        )

    # ── Voeg debiet_klasse_extreem toe aan bestaande geopackages ──
    if RUN_EXTRA_INFO:
        ExtraInfoToevoegenAllData(
            loc_koppeltabel=loc_koppeltabel,
            meas_folder=meas_folder,
            model_folder=model_folder,
            stat_cols=("abs_q95", "abs_q05"),
            threshold=5,
            new_col="P95_P05_alle_beschikbare_data",
            mode="any",
            above_operator=">=",
        )

    # ── Eindbeoordeling op basis van weggeschreven Validatie_criteria.xlsx ──
    if RUN_EINDBEOORDELING:
        validatie_xlsx = Path(model_folder) / "results" / "Validatie_criteria.xlsx"
        eindbeoordeling_xlsx = Path(model_folder) / "results" / "Eindbeoordeling_model.xlsx"
        BerekenModelEindbeoordeling(
            validatie_xlsx=validatie_xlsx,
            output_xlsx=eindbeoordeling_xlsx,
            criteria_groepen=CRITERIA_GROEPEN,
            min_groepen_voldoen=MIN_GROEPEN_VOLDOEN,
            drempel_hws=DREMPEL_HWS,
            drempel_regionaal=DREMPEL_REGIONAAL,
            hws_waterschap=HWS_WATERSCHAP,
            beoor_kleuren=BEOOR_KLEUREN,
        )

    # ── LHM 4.1-paden ────────────────────────────────────────────────────────
    lhm41_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/LHM4_1_reeksen")
    gpkg_koppellaag = cloud.joinpath(
        "Basisgegevens/resultaatvergelijking/Koppeltabel_LHM4_1/Koppeltabel_met_toegevoegde_data_v28_04_2026.gpkg"
    )
    lhm41_layer_naam = "koppeling_lhm4_1_reeksen"
    if RUN_LHM41:
        cloud.synchronize([lhm41_folder, gpkg_koppellaag])
        AnalyseLHM41Vergelijking(
            gpkg_koppellaag=gpkg_koppellaag,
            layer_naam=lhm41_layer_naam,
            model_folder=model_folder,
            meas_folder=meas_folder,
            lhm41_folder=lhm41_folder,
            criteria_lhm41=CRITERIA_LHM41,
            LHM41_TEKEN_CORRECTIE=LHM41_TEKEN_CORRECTIE,
        )

    # ── HTML-viewer genereren op basis van weggeschreven geopackages ──
    if RUN_HTML_VIEWER:
        CreateHTMLViewer(
            model_folder=model_folder,
            model_gpkg=model_gpkg_in_html,
            include_lhm41=include_lhm41_html,
            include_fractie=include_fractie_html,
        )
        # CreateHTMLViewer(model_folder=model_folder, model_gpkg=None, include_lhm41=True)

    # ── Upload resultaten naar de cloud ──
    #!TODO: nog niet getest
    # if RUN_UPLOAD:
    #     results = Path(model_folder) / "results"
    #     for gpkg in [
    #         "Validatie_resultaten_all.gpkg",
    #         "Validatie_resultaten_dec_all.gpkg",
    #         "Validatie_resultaten.gpkg",
    #         "Validatie_resultaten_dec.gpkg",
    #     ]:
    #         gpkg_path = results / gpkg
    #         if gpkg_path.exists():
    #             cloud.upload_file(gpkg_path)
    #     for xlsx in ["Validatie_criteria.xlsx", "Eindbeoordeling_model.xlsx"]:
    #         xlsx_path = results / xlsx
    #         if xlsx_path.exists():
    #             cloud.upload_file(xlsx_path)
    #     figures_path = results / "figures"
    #     if figures_path.exists():
    #         cloud.upload_content(figures_path)
    #     html_path = results / "Validatieresultaten_HTML"
    #     if html_path.exists():
    #         cloud.upload_content(html_path)

# %%
