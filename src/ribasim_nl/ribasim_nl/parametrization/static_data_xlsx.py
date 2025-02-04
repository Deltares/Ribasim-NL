import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border
from pydantic import BaseModel, ConfigDict

from ribasim_nl import Model
from ribasim_nl.case_conversions import pascal_to_snake_case
from ribasim_nl.parametrization.empty_table import empty_table_df

description = [
    {"sheet": "defaults", "beschrijving": "default settings bij geen-data"},
    {"sheet": "Outlet & Pump", "beschrijving": "gemodelleerde instellingen bij RIBASIM Pump en Outlet nodes"},
    {"kolom": "flow_rate", "beschrijving": "gemodelleerde capaciteit (m3/s) van het kunstwerk"},
    {
        "kolom": "code",
        "beschrijving": "code waterbeheerder van het object",
    },
    {
        "kolom": "name",
        "beschrijving": "naam van het object",
    },
    {
        "kolom": "flow_rate_mm_per_day",
        "beschrijving": "gemodelleerde capaciteit (mm/dag) van het kunstwerk t.o.v. voor het bovenstrooms/benedenstroomse (uitlaat/inlaat) stroomgebied",
    },
    {
        "kolom": "categorie",
        "beschrijving": "categorie van het kunstwerk. Is de logische link tussen de Pump/Outlet en defaults table.",
    },
    {
        "kolom": "opmerking_waterbeheerder",
        "beschrijving": "een opmerking van de waterbeheerder ter duiding van de ingevulde waarde(n)",
    },
    {"kolom": "upstream_level_offset", "beschrijving": "afslagpunt (meter) boven bovenstrooms streefpeil"},
    {"kolom": "downstream_level_offset", "beschrijving": "afslagpunt (meter) onder benedenstrooms streefpeil"},
]

defaults = {
    "Afvoergemaal": {
        "upstream_level_offset": 0.0,
        "downstream_level_offset": 0.2,
        "flow_rate": pd.NA,
        "flow_rate_mm_per_day": 15,
        "function": "outlet",
    },
    "Aanvoergemaal": {
        "upstream_level_offset": 0.2,
        "downstream_level_offset": 0.0,
        "flow_rate": pd.NA,
        "flow_rate_mm_per_day": 4,
        "function": "inlet",
    },
    "Uitlaat": {
        "upstream_level_offset": 0.0,
        "downstream_level_offset": 0.3,
        "flow_rate": pd.NA,
        "flow_rate_mm_per_day": 50,
        "function": "outlet",
    },
    "Inlaat": {
        "upstream_level_offset": 0.2,
        "downstream_level_offset": 0.0,
        "flow_rate": pd.NA,
        "flow_rate_mm_per_day": 4,
        "function": "inlet",
    },
}


class StaticData(BaseModel):
    xlsx_path: os.PathLike
    model: Model
    default_dict: dict = defaults
    description_list: list = description
    outlet: pd.DataFrame | None = None
    pump: pd.DataFrame | None = None

    model_config = ConfigDict(arbitrary_types_allowed=True)

    def __post_init__(self):
        self.xlsx_path = Path(self.xlsx_path)

    @property
    def defaults(self):
        df = pd.DataFrame.from_dict(self.default_dict, orient="index")
        df.index.name = "categorie"
        return df

    @property
    def description(self):
        df = pd.DataFrame(self.description_list)
        df.sort_values(by=["sheet", "kolom"], inplace=True)
        return df[["sheet", "kolom", "beschrijving"]]

    def reset_data_frame(self, node_type):
        columns = ["node_id", "name", "code", "flow_rate", "min_upstream_level", "max_downstream_level"]
        extra_columns = ["categorie", "opmerking_waterbeheerder"]
        df = empty_table_df(
            model=self.model,
            table_type="Static",
            node_type=node_type,
            meta_columns=["meta_code_waterbeheerder", "name"],
        )
        df.rename(columns={"meta_code_waterbeheerder": "code"}, inplace=True)
        if node_type == "Pump":
            df["categorie"] = "Afvoergemaal"
        if node_type == "Outlet":
            df["categorie"] = "Uitlaat"
        df["opmerking_waterbeheerder"] = ""
        df = df[columns + extra_columns]
        setattr(self, pascal_to_snake_case(node_type), df)
        return getattr(self, pascal_to_snake_case(node_type))

    def add_series(self, node_type, series):
        col = series.index.name

        # get dataframe
        df = getattr(self, pascal_to_snake_case(node_type))
        if df is None:
            df = self.reset_data_frame(node_type=node_type)

        # sanitize series to existing codes in node_type
        series = series[series.index.isin(df[col])]

        # mask table where row in series and set data
        mask = df[col].isin(series.index)
        df.loc[mask, series.name] = series[df[mask][col]].to_numpy()

    def write(self):
        # remove exel if exists
        if self.xlsx_path.exists():
            self.xlsx_path.unlink()
        else:
            self.xlsx_path.parent.mkdir(exist_ok=True, parents=True)

        # write sheets
        self.description.to_excel(self.xlsx_path, sheet_name="beschrijving", index=False)
        with pd.ExcelWriter(self.xlsx_path, mode="a", if_sheet_exists="replace") as xlsx_writer:
            self.defaults.to_excel(xlsx_writer, sheet_name="defaults")
            for node_type in ["Pump", "Outlet"]:
                df = getattr(self, pascal_to_snake_case(node_type))
                if df is None:
                    df = self.reset_data_frame(node_type=node_type)
                df.to_excel(xlsx_writer, sheet_name=node_type, index=False)

        # fix formatting
        wb = load_workbook(self.xlsx_path)
        for sheet_name in pd.ExcelFile(self.xlsx_path).sheet_names:
            ws = wb[sheet_name]

            # freeze panes
            ws.freeze_panes = "A2"

            # add filters
            ws.auto_filter.ref = ws.dimensions

            # Remove cell borders
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border()  # Removes borders

            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Get column letter (e.g., 'A', 'B', etc.)

                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding

            # Save the modified file
            wb.save(self.xlsx_path)
        wb.close()
