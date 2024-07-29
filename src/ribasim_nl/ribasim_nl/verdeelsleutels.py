from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame, Series
from ribasim.nodes import discrete_control

from ribasim_nl.model import add_control_node_to_network


def read_verdeelsleutel(file_path: Path) -> DataFrame:
    """Concat sheets in a verdeelsleutel.xlsx to 1 pandas dataframe."""
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    return pd.concat([pd.read_excel(file_path, sheet_name=i) for i in sheet_names])


def verdeelsleutel_to_fractions(
    verdeelsleutel_df: DataFrame, node_index: Series, keys: int = 2
) -> DataFrame:
    df = pd.concat(
        [
            verdeelsleutel_df[
                [f"locatie_benedenstrooms_{i}", f"fractie_{i}", "beschrijving"]
            ].rename(
                columns={
                    f"locatie_benedenstrooms_{i}": "locatie_benedenstrooms",
                    f"fractie_{i}": "fraction",
                    "beschrijving": "control_state",
                }
            )
            for i in range(1, keys + 1)
        ]
    )

    for code, node_id in zip(node_index.index, node_index):
        df.loc[
            df.locatie_benedenstrooms.str.lower() == code.lower(), "node_id"
        ] = node_id

    df.loc[:, "fraction"] = df["fraction"].round(3)
    return df[["node_id", "fraction", "control_state"]]


def verdeelsleutel_to_control(
    verdeelsleutel_df,
    model,
    offset_node_id=None,
    code_waterbeheerder=None,
    waterbeheerder=None,
):
    keys = verdeelsleutel_df.locatie_bovenstrooms.unique()
    frac_nodes = []
    if len(keys) != 1:
        raise ValueError(
            f"number of keys in verdeelsleutel != 1: {verdeelsleutel_df.locatie_bovenstrooms.unique()}"
        )
    else:
        listen_node_id = (
            model.node_table()
            .df[model.node_table().df["meta_code_waterbeheerder"] == keys[0]]
            .iloc[0]
            .node_id
        )

    # get frac-nodes
    for (loc1, loc2), df in verdeelsleutel_df.groupby(
        ["locatie_benedenstrooms_1", "locatie_benedenstrooms_2"]
    ):
        frac_nodes += [
            model.node_table()
            .df[
                (model.node_table().df["node_type"] == "FractionalFlow")
                & (
                    model.node_table().df["meta_code_waterbeheerder"].str.lower()
                    == i.lower()
                )
            ]
            .iloc[0]
            .node_id
            for i in [loc1, loc2]
        ]

    ctrl_node = add_control_node_to_network(
        model,
        frac_nodes,
        offset=100,
        offset_node_id=offset_node_id,
        meta_code_waterbeheerder=code_waterbeheerder,
        meta_waterbeheerder=waterbeheerder,
    )

    # add descrete control
    condition_df = df[["ondergrens_waarde", "beschrijving"]].rename(
        columns={
            "ondergrens_waarde": "greater_than",
            "beschrijving": "meta_beschrijving",
        }
    )
    condition_df["node_id"] = ctrl_node.node_id
    condition_df["listen_node_id"] = listen_node_id
    condition_df["variable"] = "flow_rate"

    logic_df = df[["beschrijving"]].rename(columns={"beschrijving": "control_state"})
    logic_df["truth_state"] = [
        "".join(["T"] * i + ["F"] * len(df))[0 : len(df)] for i in range(len(df))
    ]
    logic_df["node_id"] = ctrl_node.node_id

    listen_node_type = (
        model.node_table().df.set_index("node_id").at[listen_node_id, "node_type"]
    )
    model.discrete_control.add(
        ctrl_node,
        [
            discrete_control.Condition(
                listen_node_id=listen_node_id,
                listen_node_type=listen_node_type,
                variable="flow_rate",
                greater_than=condition_df["greater_than"].to_list(),
            ),
            discrete_control.Logic(
                truth_state=logic_df.truth_state.to_list(),
                control_state=logic_df.control_state.to_list(),
            ),
        ],
    )

    for frac_node_id in frac_nodes:
        model.edge.add(
            model.discrete_control[ctrl_node.node_id],
            model.fractional_flow[frac_node_id],
        )

    return model
