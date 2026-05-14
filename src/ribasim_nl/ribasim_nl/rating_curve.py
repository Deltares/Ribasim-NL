from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame, Series


def read_rating_curve(file_path: Path, node_index: Series) -> DataFrame:
    """Concat sheets in a verdeelsleutel.xlsx to 1 pandas dataframe."""
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    dfs = []
    for sheet_name in sheet_names:
        if sheet_name != "disclaimer":
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df["code_waterbeheerder"] = sheet_name
            df["node_id"] = node_index.loc[sheet_name]
            dfs += [df]

    return pd.concat(dfs)[["node_id", "level", "flow_rate", "code_waterbeheerder"]]


def _prepare_qh_series(series: Series) -> Series:
    if series.name is None:
        raise ValueError("Each Q(H) series must have a name.")
    if series.empty:
        raise ValueError(f"Q(H) series '{series.name}' is empty.")
    if series.index.has_duplicates:
        raise ValueError(f"Q(H) series '{series.name}' contains duplicate H values.")

    qh_series = series.sort_index()
    if not qh_series.index.is_monotonic_increasing:
        raise ValueError(f"Q(H) series '{series.name}' could not be sorted on H.")

    return qh_series.astype(float)


def _get_common_h_step(qh_series_list: list[Series]) -> float:
    h_values = sorted({float(value) for series in qh_series_list for value in series.index.to_numpy(dtype=float)})
    diffs = pd.Series(h_values).diff().dropna()
    positive_diffs = diffs[diffs > 0]

    if positive_diffs.empty:
        return 1.0
    return float(positive_diffs.min())


def flow_distribution_by_level(qh_series_list: list[Series]) -> DataFrame:
    """Compute discharge fractions for multiple Q(H) relationships on a common H grid.

    The H-grid follows the series with the largest H-range. All series are linearly
    interpolated to that grid. Outside a series' H-range, its minimum or maximum Q
    value is used.
    """
    if not qh_series_list:
        raise ValueError("At least one Q(H) series is required.")

    prepared_series = [_prepare_qh_series(series) for series in qh_series_list]

    duplicate_names = pd.Index([series.name for series in prepared_series]).duplicated()
    if duplicate_names.any():
        raise ValueError("Each Q(H) series must have a unique name.")

    leader = max(
        prepared_series,
        key=lambda series: (float(series.index.max()) - float(series.index.min()), len(series)),
    )

    h_step = _get_common_h_step(prepared_series)
    h_min = float(leader.index.min())
    h_max = float(leader.index.max())
    h_values = pd.Index(pd.Series([h_min + i * h_step for i in range(int((h_max - h_min) / h_step) + 1)]), dtype=float)

    if h_values.empty or h_values[-1] < h_max:
        h_values = h_values.append(pd.Index([h_max], dtype=float))
    elif h_values[-1] > h_max:
        h_values = pd.Index([*h_values[:-1], h_max], dtype=float)

    q_df = DataFrame(index=h_values)
    for series in prepared_series:
        q_df[series.name] = pd.Series(
            np.interp(
                h_values.to_numpy(dtype=float),
                series.index.to_numpy(dtype=float),
                series.to_numpy(dtype=float),
                left=float(series.iloc[0]),
                right=float(series.iloc[-1]),
            ),
            index=h_values,
        )

    q_sum = q_df.sum(axis=1)
    fractions_df = q_df.div(q_sum.where(q_sum != 0), axis=0).fillna(0.0)
    fractions_df.index.name = leader.index.name or "H"
    return fractions_df
