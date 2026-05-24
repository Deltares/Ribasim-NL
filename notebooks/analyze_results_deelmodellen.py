# %%
"""Validatieanalyse deel modellen."""

from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl.styles import PatternFill
from ribasim_nl.analyse_results import (
    AnalyseLHM41Vergelijking,
    BerekenModelEindbeoordeling,
    CompareOutputMeasurements,
    ExtraInfoToevoegenAllData,
)
from ribasim_nl.html_viewer import CreateHTMLViewer
from shapely import wkt

from ribasim_nl import CloudStorage, Model


def fix_koppeltabel_new_link_id(
    koppeltabel_path: str | Path,
    model_path: str | Path,
    output_path: str | Path | None = None,
    updates_path: str | Path | None = None,
    relevant_authorities: set[str] | None = None,
    max_distance_m: float | None = 5000.0,
) -> tuple[Path, Path, pd.DataFrame]:
    koppeltabel_path = Path(koppeltabel_path)
    model_path = Path(model_path)

    if output_path is None:
        output_path = koppeltabel_path.with_name(f"{koppeltabel_path.stem}_post_fixen.xlsx")
    else:
        output_path = Path(output_path)

    if updates_path is None:
        updates_path = koppeltabel_path.with_name(f"{koppeltabel_path.stem}_post_fixen_updates.csv")
    else:
        updates_path = Path(updates_path)

    df = pd.read_excel(koppeltabel_path)

    model = Model.read(model_path)
    links = gpd.GeoDataFrame(model.link.df.copy(), geometry="geometry", crs=model.link.df.crs)
    links = links[links["link_type"] == "flow"].copy().reset_index()

    points = gpd.GeoDataFrame(df.copy(), geometry=df["geometry"].apply(wkt.loads), crs=links.crs)

    nearest = (
        gpd.sjoin_nearest(
            points[["Waterschap", "MeetreeksC", "Aan/Af", "new_link_id", "geometry"]].copy(),
            links[["link_id", "from_node_id", "to_node_id", "geometry"]].copy(),
            how="left",
            distance_col="distance_to_link_m",
            lsuffix="pt",
            rsuffix="lnk",
        )
        .reset_index()
        .rename(columns={"index": "row_index"})
    )

    nearest = nearest.sort_values(["row_index", "distance_to_link_m", "link_id"]).drop_duplicates("row_index")

    updated_df = df.copy()
    changes = []

    for row in nearest.itertuples(index=False):
        row_index = int(row.row_index)
        waterschap = updated_df.at[row_index, "Waterschap"]

        if relevant_authorities is not None and waterschap not in relevant_authorities:
            continue

        if pd.isna(row.link_id):
            continue

        if max_distance_m is not None and float(row.distance_to_link_m) > max_distance_m:
            continue

        old_value = updated_df.at[row_index, "new_link_id"]
        new_value = str([int(row.link_id)])

        if old_value == new_value:
            continue

        updated_df.at[row_index, "new_link_id"] = new_value

        changes.append(
            {
                "row_index": row_index,
                "Waterschap": waterschap,
                "MeetreeksC": updated_df.at[row_index, "MeetreeksC"],
                "Aan/Af": updated_df.at[row_index, "Aan/Af"],
                "geometry": updated_df.at[row_index, "geometry"],
                "old_new_link_id": old_value,
                "new_new_link_id": new_value,
                "matched_link_id": int(row.link_id),
                "from_node_id": int(row.from_node_id),
                "to_node_id": int(row.to_node_id),
                "distance_to_link_m": float(row.distance_to_link_m),
            }
        )

    changes_df = pd.DataFrame(changes)
    updated_df.to_excel(output_path, index=False)
    changes_df.to_csv(updates_path, index=False)

    return output_path, updates_path, changes_df


# %%

########################################################################################################################################
# Implementatie lokaal voor testen
cloud = CloudStorage()
base_koppeltabel = cloud.joinpath("Basisgegevens/resultaatvergelijking/koppeltabel_2026")

loc_koppeltabel = (
    base_koppeltabel
    / "Transformed_koppeltabel_versie_Samenwerkdag_26052026_Feedback_Verwerkt_HydroLogic_rev_D2Hydro_230526.xlsx"
)
loc_specifieke_bewerking = base_koppeltabel / "Specifiek_bewerking_versieSamenwerkdag_26052026_revD2Hydro_230526.xlsx"

meas_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/meetreeksen_2026")

model_folder = Path(r"d:/repositories/Ribasim-NL/data/Rijkswaterstaat/modellen/lhm_sub_models/AAM-Limburg-RWS_coupled")
toml_naam = "AAM-Limburg-RWS_coupled.toml"

# synchronize paths
cloud.synchronize(
    [
        # loc_koppeltabel,
        # loc_specifieke_bewerking,
        meas_folder,
    ]
)

########################################################################################################################################

# ── Toetsingscriteria en drempelwaarden ──────────────────────────────────
# Richtwaarden: Goed = enige klasse (geen Matig); anders Onvoldoende.
# NSE > 0.4, KGE > 0.5, |Bias| < 15%, P10/P90 binnen ±30%, P25/P75 binnen ±15%.
CRITERIA_GRENZEN = {
    "Bias": {"Goed": 15.0},
    "NSE": {"Goed": 0.4},
    "KGE": {"Goed": 0.5},
    "P10": {"Goed": 30.0},
    "P25": {"Goed": 15.0},
    "P75": {"Goed": 15.0},
    "P90": {"Goed": 30.0},
}
BEOOR_KLEUREN = {
    "Goed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Matig": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Onvoldoende": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "n.v.t.": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
# 6 kwantitatieve criteria; model voldoet als ≥ MIN_GROEPEN_VOLDOEN ervan "Goed" zijn.
# NSE én KGE tellen samen als één criterium (tenminste één moet voldoen).
CRITERIA_GROEPEN = {
    "Bias": {"kolommen": ["Beoor_Bias_dec"], "min_goed": 1},
    "NSE_KGE": {"kolommen": ["Beoor_NSE_dec", "Beoor_KGE_dec"], "min_goed": 1},
    "P10": {"kolommen": ["Beoor_P10_dec"], "min_goed": 1},
    "P25": {"kolommen": ["Beoor_P25_dec"], "min_goed": 1},
    "P75": {"kolommen": ["Beoor_P75_dec"], "min_goed": 1},
    "P90": {"kolommen": ["Beoor_P90_dec"], "min_goed": 1},
}
MIN_GROEPEN_VOLDOEN = 3  # ≥ 3 van de 6 criteria Goed
DREMPEL_HWS = 75.0  # HWS: 75% van locaties moet voldoen
DREMPEL_REGIONAAL = 50.0  # Regionaal: 50% van locaties moet voldoen
HWS_WATERSCHAP = "Rijkswaterstaat"
# Absolute vloer voor Bias en percentielen: bij |sim - obs| < drempel altijd Goed
ABS_DREMPEL_M3S = 0.15

# ── LHM 4.1-toetsingscriteria ────────────────────────────────────────────
CRITERIA_LHM41 = {
    ("Afvoer", ">=5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumJaar", 25, 80, "Cum. jaarafvoer dif. < 25%"),
    ],
    ("Afvoer", "<5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumJaar", 33, 80, "Cum. jaarafvoer dif. < 33%"),
    ],
    ("Aanvoer", ">=5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumZomer", 25, 80, "Cum. zomer aanvoer dif. < 25%"),
    ],
    ("Aanvoer", "<5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumZomer", 50, 80, "Cum. zomer aanvoer dif. < 50%"),
    ],
}
# Tekencorrectie per LHM 4.1 CSV-bestand (-1 als tekenconventie verschilt van Ribasim/meting)
LHM41_TEKEN_CORRECTIE: dict[str, float] = {
    "Aanvoer Gemaal Winsemius_LHM_reeks.csv": -1.0,
}

RUN_COMPARE = True
RUN_EXTRA_INFO = True
RUN_EINDBEOORDELING = True
RUN_LHM41 = True

RUN_HTML_VIEWER = True
# Dit gaat nog niet ivm hoeveelheid links, polygons en nodes in modellen
model_gpkg_in_html = None
# Als we ook resultaten vanuit LHM4.1 kunnen we deze laag wel of niet tonen in html viewer
#!!!TODO: Draai hiervoor eerst notebooks/observations/analyse_lhm41_vergelijking
include_lhm41_html = RUN_LHM41
# Als we ook fracties kunnen plotten en tonen in html viewer
include_fractie_html = False

RUN_UPLOAD = False

# %%

# EXCLUDE_MEETREEKS = ["Megen dorp Hoofdkranen", "Weesp West Hoofdkranen", "genemuiden, de ketting Hoofdkranen",
#                      "Panheel Hoofdkranen", "Almen vechtstromen", "Bovensluis", "Brienenoord (kilometer 996.5)",
#                       "Keizersveer", "Maassluis", "Pannerden", "Puttershoek", "Venlo", "Eem Eemdijk"]

# ── Verwerk meetreeksen, bereken statistieken, schrijf figuren/geopackages/Excel ──
if RUN_COMPARE:
    CompareOutputMeasurements(
        loc_koppeltabel=loc_koppeltabel,
        loc_specifics=loc_specifieke_bewerking,
        meas_folder=meas_folder,
        model_folder=model_folder,
        toml_file=toml_naam,
        filetype="flow",
        criteria_grenzen=CRITERIA_GRENZEN,
        beoor_kleuren=BEOOR_KLEUREN,
        abs_drempel=ABS_DREMPEL_M3S,
        exclude_meetreeks=None,
        save_results_combined=True,
        output_is_feather=False,
        output_is_nc=True,
        resample_to_daily=True,
    )
else:
    print(
        "LET OP: RUN_COMPARE is uitgeschakeld. De overige stappen (RUN_EXTRA_INFO, "
        "RUN_EINDBEOORDELING, RUN_LHM41, RUN_HTML_VIEWER) vereisen dat "
        "CompareOutputMeasurements al eerder succesvol is gedraaid."
    )

# ── Voeg debiet_klasse_extreem toe aan bestaande geopackages ──
if RUN_EXTRA_INFO:
    ExtraInfoToevoegenAllData(
        loc_koppeltabel=loc_koppeltabel,
        meas_folder=meas_folder,
        model_folder=model_folder,
        stat_cols=("abs_q95", "abs_q05"),
        threshold=5,
        new_col="P95_P05_alle_beschikbare_data",
        mode="any",
        above_operator=">=",
    )

# ── Eindbeoordeling op basis van weggeschreven Validatie_criteria.xlsx ──
if RUN_EINDBEOORDELING:
    validatie_xlsx = Path(model_folder) / "results" / "Validatie_criteria.xlsx"
    eindbeoordeling_xlsx = Path(model_folder) / "results" / "Eindbeoordeling_model.xlsx"
    BerekenModelEindbeoordeling(
        validatie_xlsx=validatie_xlsx,
        output_xlsx=eindbeoordeling_xlsx,
        criteria_groepen=CRITERIA_GROEPEN,
        min_groepen_voldoen=MIN_GROEPEN_VOLDOEN,
        drempel_hws=DREMPEL_HWS,
        drempel_regionaal=DREMPEL_REGIONAAL,
        hws_waterschap=HWS_WATERSCHAP,
        beoor_kleuren=BEOOR_KLEUREN,
    )

# ── LHM 4.1-paden ────────────────────────────────────────────────────────
lhm41_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/LHM4_1_reeksen")
gpkg_koppellaag = cloud.joinpath(
    "Basisgegevens/resultaatvergelijking/Koppeltabel_LHM4_1/Koppeltabel_met_toegevoegde_data_v28_04_2026.gpkg"
)
lhm41_layer_naam = "koppeling_lhm4_1_reeksen"
if RUN_LHM41:
    cloud.synchronize([lhm41_folder, gpkg_koppellaag])
    AnalyseLHM41Vergelijking(
        gpkg_koppellaag=gpkg_koppellaag,
        layer_naam=lhm41_layer_naam,
        model_folder=model_folder,
        meas_folder=meas_folder,
        lhm41_folder=lhm41_folder,
        criteria_lhm41=CRITERIA_LHM41,
        LHM41_TEKEN_CORRECTIE=LHM41_TEKEN_CORRECTIE,
    )

# ── HTML-viewer genereren op basis van weggeschreven geopackages ──
if RUN_HTML_VIEWER:
    CreateHTMLViewer(
        model_folder=model_folder,
        model_gpkg=model_gpkg_in_html,
        include_lhm41=include_lhm41_html,
        include_fractie=include_fractie_html,
    )
    # CreateHTMLViewer(model_folder=model_folder, model_gpkg=None, include_lhm41=True)

# ── Upload resultaten naar de cloud ──
#!TODO: nog niet getest
# if RUN_UPLOAD:
#     results = Path(model_folder) / "results"
#     for gpkg in [
#         "Validatie_resultaten_all.gpkg",
#         "Validatie_resultaten_dec_all.gpkg",
#         "Validatie_resultaten.gpkg",
#         "Validatie_resultaten_dec.gpkg",
#     ]:
#         gpkg_path = results / gpkg
#         if gpkg_path.exists():
#             cloud.upload_file(gpkg_path)
#     for xlsx in ["Validatie_criteria.xlsx", "Eindbeoordeling_model.xlsx"]:
#         xlsx_path = results / xlsx
#         if xlsx_path.exists():
#             cloud.upload_file(xlsx_path)
#     figures_path = results / "figures"
#     if figures_path.exists():
#         cloud.upload_content(figures_path)
#     html_path = results / "Validatieresultaten_HTML"
#     if html_path.exists():
#         cloud.upload_content(html_path)

# %%
