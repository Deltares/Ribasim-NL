# %%
"""Vergelijking Ribasim LHM 5.x vs. LHM 4.1 op decade-basis.

Werkt met een selectie van meetlocaties waarvoor ook LHM 4.1 decade-uitvoer
beschikbaar is. Hergebruikt functies uit analyse_results.py.

Workflow:
1. Laad de koppellaag (geopackage) met de locatie-selectie en LHM 4.1 CSV-namen.
2. Lees per locatie de weggeschreven decade-tijdreeks (Ribasim + meting).
3. Groepeer op LHM_4_1_csv: sommeer Ribasim- en meetwaarden voor rijen die
   dezelfde LHM 4.1 CSV delen.
4. Lees de LHM 4.1 decade-CSV en merge op datum.
5. Bereken statistieken voor Ribasim vs. meting (per locatie) en
   LHM 4.1 vs. meting (per groep).
6. Evalueer toetsingscriteria en schrijf Validatie_criteria_lhm4_1.xlsx.
"""

from collections import defaultdict
from pathlib import Path

import geopandas as gpd
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from ribasim_nl.analyse_results import (
    GetStatisticsPerPeriod,
    LoadMeasurements,
    _sanitize_filename,
)

from ribasim_nl import CloudStorage

# %%
# ── Toetsingsdrempels ────────────────────────────────────────────────────────
# Elke tuple: (type_stat, drempel, pct_vereist, label)
# type_stat: "NSE" | "CumJaar" | "CumZomer"
# drempel:   voor NSE geldt stat > drempel; voor Cum geldt afwijking [%] < drempel
# pct_vereist: minimaal % locaties dat moet voldoen

CRITERIA_LHM41 = {
    ("Afvoer", ">=5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumJaar", 25, 80, "Cum. jaarafvoer dif. < 25%"),
    ],
    ("Afvoer", "<5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumJaar", 33, 80, "Cum. jaarafvoer dif. < 33%"),
    ],
    ("Aanvoer", ">=5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumZomer", 25, 80, "Cum. zomer aanvoer dif. < 25%"),
    ],
    ("Aanvoer", "<5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumZomer", 50, 80, "Cum. zomer aanvoer dif. < 50%"),
    ],
}

_AANAF_NAAR_CATEGORIE = {
    "Afvoer": "Afvoer",
    "Aanvoer": "Aanvoer",
    "Aan&Af": "Afvoer",  # behandeld als afvoer voor toetsing
}

# Tekencorrectie per LHM 4.1 CSV-bestand.
# Gebruik -1 als de LHM-reeks een andere tekenconventie heeft dan Ribasim en de meting.
# Voeg nieuwe gevallen toe als dict-item: "bestandsnaam.csv": vermenigvuldigingsfactor
LHM41_TEKEN_CORRECTIE: dict[str, float] = {
    "Aanvoer Gemaal Winsemius_LHM_reeks.csv": -1.0,
}

BEOOR_KLEUREN = {
    "Goed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Matig": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Onvoldoende": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "n.v.t.": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


# ── Hulpfuncties ──────────────────────────────────────────────────────────────


def _laad_decade_tijdreeks(ts_root: str, waterschap: str, fig_name_clean: str) -> pd.DataFrame | None:
    """Leest de decade-sheet uit de weggeschreven tijdreeks-Excel.

    Retourneert DataFrame met kolommen: Datum, meting, Ribasim.
    """
    path = Path(ts_root) / waterschap / f"{fig_name_clean}.xlsx"
    if not path.exists():
        print(f"  Tijdreeks niet gevonden: {path}")
        return None
    df = pd.read_excel(path, sheet_name="Decade")
    df = df.iloc[:, :3].copy()
    df.columns = ["Datum", "meting", "Ribasim"]
    df["Datum"] = pd.to_datetime(df["Datum"])
    df["meting"] = pd.to_numeric(df["meting"], errors="coerce")
    df["Ribasim"] = pd.to_numeric(df["Ribasim"], errors="coerce")
    return df


def _laad_lhm41_csv(lhm41_folder: str, csv_naam: str) -> pd.DataFrame | None:
    """Leest een LHM 4.1 decade-CSV.

    Verwacht: kolom 1 = Startdate (YYYY-MM-DD), kolom 2 = waarde [m3/s of unit].
    Retourneert DataFrame met kolommen: Datum, LHM41.
    """
    path = Path(lhm41_folder) / csv_naam
    if not path.exists():
        print(f"  LHM 4.1 CSV niet gevonden: {path}")
        return None
    df = pd.read_csv(path, sep=None, engine="python")
    df = df.iloc[:, :2].copy()
    df.columns = ["Datum", "LHM41"]
    df["Datum"] = pd.to_datetime(df["Datum"])
    df["LHM41"] = pd.to_numeric(df["LHM41"], errors="coerce")
    factor = LHM41_TEKEN_CORRECTIE.get(csv_naam)
    if factor is not None:
        df["LHM41"] *= factor
        print(f"  Tekencorrectie toegepast op {csv_naam} (factor: {factor})")
    return df


def _bereken_p95_klasse(
    measurements: dict,
    meetreeks_c: str,
    aan_af: str,
    threshold: float = 5.0,
) -> str:
    """Bepaalt de P95_P05-klasse op basis van de volledige dagmeetreeks."""
    meas_df = measurements["afvoer_dag"] if aan_af in ("Afvoer", "Aan&Af") else measurements["aanvoer_dag"]
    if meetreeks_c not in meas_df.columns:
        return "geen_data"
    serie = meas_df[meetreeks_c].dropna()
    if len(serie) == 0:
        return "geen_data"
    q05 = serie.quantile(0.05)
    q95 = serie.quantile(0.95)
    return ">=5" if (abs(q05) >= threshold or abs(q95) >= threshold) else "<5"


def _bereken_p95_klasse_som(
    measurements: dict,
    locs: list,
    threshold: float = 5.0,
) -> str:
    """P95-klasse op basis van de gesommeerde dagmeetreeksen van alle locaties in de groep.

    Alleen datums waarop ALLE reeksen data hebben worden meegenomen (intersectie),
    zodat de som niet kunstmatig laag wordt door gedeeltelijke ontbrekende data.
    """
    aan_af = locs[0]["Aan/Af"]
    meas_df = measurements["afvoer_dag"] if aan_af in ("Afvoer", "Aan&Af") else measurements["aanvoer_dag"]

    series_list = []
    for loc in locs:
        col = loc["MeetreeksC"]
        if col in meas_df.columns:
            series_list.append(meas_df[col].rename(col))

    if not series_list:
        return "geen_data"

    # pd.concat + dropna geeft intersectie: alleen rijen waar alle reeksen data hebben
    combined = pd.concat(series_list, axis=1).dropna()
    if combined.empty:
        return "geen_data"

    som = combined.sum(axis=1)
    q05 = som.quantile(0.05)
    q95 = som.quantile(0.95)
    return ">=5" if (abs(q05) >= threshold or abs(q95) >= threshold) else "<5"


def _cum_afwijking_jaarlijks(
    df: pd.DataFrame,
    col_model: str,
    col_obs: str,
    zomer: bool = False,
) -> dict:
    """Relatieve afwijking cumulatieve jaarafvoer per jaar én gemiddeld [%].

    Methodiek conform het oude LHM 4.1 toetsingsscript:
    - Per jaar wordt de cumulatieve som opgebouwd t/m de laatste dag met een
      geldige meting in dat jaar (``last_valid_index``-equivalent).
    - Afwijking per jaar = ``|1 - cum_model / cum_meting| * 100``.
    - Samenvatting = ``nanmean`` over alle jaren (geen mediaan).
    - Daarnaast de afwijking over de volledige analyseperiode als geheel.

    Bij zomer=True wordt alleen het zomerhalfjaar (april-september) meegenomen.

    Retourneert een dict met:
    - ``per_jaar``:  {jaar: afwijking [%]}
    - ``gemiddeld``: gemiddelde over alle jaren
    - ``totaal``:    afwijking over de gehele periode
    """
    leeg = {"per_jaar": {}, "gemiddeld": np.nan, "totaal": np.nan}

    d = df.copy()
    if zomer:
        d = d[d["Datum"].dt.month.between(4, 9)]
    if d.empty:
        return leeg

    d["jaar"] = d["Datum"].dt.year
    afwijkingen_per_jaar = {}

    for jaar, grp in d.groupby("jaar"):
        # Alleen rijen tot en met de laatste geldige meting in dit jaar
        grp_geldig = grp.dropna(subset=[col_obs])
        if grp_geldig.empty:
            continue
        laatste_datum = grp_geldig["Datum"].max()
        grp_tot_einde = grp[grp["Datum"] <= laatste_datum]

        cum_obs = grp_tot_einde[col_obs].fillna(0).sum()
        cum_mod = grp_tot_einde[col_model].fillna(0).sum()

        if abs(cum_obs) > 0:
            afwijkingen_per_jaar[int(jaar)] = round(abs(1 - cum_mod / cum_obs) * 100, 1)

    # Gemiddelde over jaren
    waarden = list(afwijkingen_per_jaar.values())
    gemiddeld = float(np.nanmean(waarden)) if waarden else np.nan

    # Totale periode: cumulatief t/m laatste geldige meting in de gehele dataset
    d_geldig = d.dropna(subset=[col_obs])
    if not d_geldig.empty:
        laatste_datum_totaal = d_geldig["Datum"].max()
        d_tot_einde = d[d["Datum"] <= laatste_datum_totaal]
        cum_obs_tot = d_tot_einde[col_obs].fillna(0).sum()
        cum_mod_tot = d_tot_einde[col_model].fillna(0).sum()
        totaal = round(abs(1 - cum_mod_tot / cum_obs_tot) * 100, 1) if abs(cum_obs_tot) > 0 else np.nan
    else:
        totaal = np.nan

    return {"per_jaar": afwijkingen_per_jaar, "gemiddeld": gemiddeld, "totaal": totaal}


def _bereken_statistieken(
    df: pd.DataFrame,
    col_model: str,
    col_obs: str,
    categorie: str,
) -> dict:
    """Berekent statistieken voor een locatie of groep.

    NSE-sleutels:
    - ``NSE``          : NSE over de totale periode (gebruikt voor toetsing)
    - ``NSE_per_jaar`` : dict {jaar: NSE}

    Cumulatieve sleutels:
    - ``CumJaar``  / ``CumZomer``              : gemiddelde over jaren (gebruikt voor toetsing)
    - ``CumJaar_totaal`` / ``CumZomer_totaal`` : afwijking over de gehele periode
    - ``CumJaar_per_jaar`` / ``CumZomer_per_jaar`` : dict {jaar: afwijking}
    """
    df_stats = df.rename(columns={col_model: "flow_rate", col_obs: "sum", "Datum": "time"})

    # NSE totaal + per jaar via GetStatisticsPerPeriod
    stats_per_periode = GetStatisticsPerPeriod(df_stats)
    stats = dict(stats_per_periode["totaal"])  # bevat NSE, RMSE, MAE, RelBias, KGE, ...

    nse_per_jaar = {
        jaar: s["NSE"]
        for jaar, s in stats_per_periode.items()
        if jaar != "totaal" and s is not None and not np.isnan(s.get("NSE", np.nan))
    }
    stats["NSE_per_jaar"] = nse_per_jaar

    cum_jaar = _cum_afwijking_jaarlijks(df, col_model, col_obs, zomer=False)
    cum_zomer = _cum_afwijking_jaarlijks(df, col_model, col_obs, zomer=True)

    stats["CumJaar"] = cum_jaar["gemiddeld"]
    stats["CumJaar_totaal"] = cum_jaar["totaal"]
    stats["CumJaar_per_jaar"] = cum_jaar["per_jaar"]
    stats["CumZomer"] = cum_zomer["gemiddeld"]
    stats["CumZomer_totaal"] = cum_zomer["totaal"]
    stats["CumZomer_per_jaar"] = cum_zomer["per_jaar"]

    return stats


def _toets_locaties(
    stats_lijst: list[dict],
    categorie: str,
    p95_klasse: str,
) -> list[dict]:
    """Evalueert de toetsingscriteria voor een reeks locatie-statistieken."""
    criteria = CRITERIA_LHM41.get((categorie, p95_klasse), [])
    rijen = []
    for type_stat, drempel, pct_vereist, label in criteria:
        waarden = [s[type_stat] for s in stats_lijst if not np.isnan(s.get(type_stat, np.nan))]
        n_totaal = len(waarden)
        if n_totaal == 0:
            rijen.append(
                {
                    "Criterium": label,
                    "N": 0,
                    "N_voldoet": 0,
                    "Pct_voldoet": np.nan,
                    "Pct_vereist": pct_vereist,
                    "Beoordeling": "n.v.t.",
                }
            )
            continue
        n_voldoet = sum(v > drempel for v in waarden) if type_stat == "NSE" else sum(v < drempel for v in waarden)
        pct = n_voldoet / n_totaal * 100
        rijen.append(
            {
                "Criterium": label,
                "N": n_totaal,
                "N_voldoet": n_voldoet,
                "Pct_voldoet": round(pct, 1),
                "Pct_vereist": pct_vereist,
                "Beoordeling": "Goed" if pct >= pct_vereist else "Onvoldoende",
            }
        )
    return rijen


def _stats_regels(stats_rib: dict, stats_lhm: dict, categorie: str, p95_klasse: str) -> list[dict]:
    """Geeft een lijst van regelobjecten terug voor gekleurde rendering in de legenda.

    Elk object heeft: {"text": str, "color": str | None, "bold": bool, "indent": bool}
    color=None betekent standaard zwart.
    """
    criteria = CRITERIA_LHM41.get((categorie, p95_klasse), [])

    def _beoordeling(waarde: float, type_stat: str, drempel: float) -> tuple[str, str]:
        if np.isnan(waarde):
            return "n.b.", "gray"
        voldoet = waarde > drempel if type_stat == "NSE" else waarde < drempel
        return ("Voldoet", "#2a7a2a") if voldoet else ("Voldoet niet", "#c0392b")

    regels = []
    regels.append(
        {"text": f"Categorie: {categorie}  |  P95: {p95_klasse} m³/s", "color": None, "bold": True, "indent": False}
    )
    regels.append({"text": "", "color": None, "bold": False, "indent": False})

    for type_stat, drempel, pct_vereist, label in criteria:
        rib_val = stats_rib.get(type_stat, np.nan)
        lhm_val = stats_lhm.get(type_stat, np.nan)
        rib_str = f"{rib_val:.1f}" if not np.isnan(rib_val) else "n.b."
        lhm_str = f"{lhm_val:.1f}" if not np.isnan(lhm_val) else "n.b."
        rib_label, rib_kleur = _beoordeling(rib_val, type_stat, drempel)
        lhm_label, lhm_kleur = _beoordeling(lhm_val, type_stat, drempel)

        regels.append({"text": f"{label}  (norm ≥{pct_vereist}% loc.)", "color": None, "bold": False, "indent": False})
        regels.append(
            {
                "text": f"  Ribasim:  {rib_str}",
                "color": None,
                "bold": False,
                "indent": True,
                "suffix": rib_label,
                "suffix_color": rib_kleur,
            }
        )
        regels.append(
            {
                "text": f"  LHM 4.1:  {lhm_str}",
                "color": None,
                "bold": False,
                "indent": True,
                "suffix": lhm_label,
                "suffix_color": lhm_kleur,
            }
        )
        regels.append({"text": "", "color": None, "bold": False, "indent": False})

    # Extra statistieken (niet in criteria)
    for sleutel, label in [("NSE", "NSE totaal")]:
        if not any(c[0] == sleutel for c in criteria):
            rib_val = stats_rib.get(sleutel, np.nan)
            lhm_val = stats_lhm.get(sleutel, np.nan)
            rib_str = f"{rib_val:.2f}" if not np.isnan(rib_val) else "n.b."
            lhm_str = f"{lhm_val:.2f}" if not np.isnan(lhm_val) else "n.b."
            regels.append({"text": label, "color": None, "bold": False, "indent": False})
            regels.append(
                {
                    "text": f"  Ribasim: {rib_str}   LHM 4.1: {lhm_str}",
                    "color": "dimgray",
                    "bold": False,
                    "indent": True,
                }
            )
            regels.append({"text": "", "color": None, "bold": False, "indent": False})

    return regels


def _plot_vergelijking(
    df_merged: pd.DataFrame,
    titel: str,
    output_folder: str,
    fig_naam: str,
    stats_rib: dict | None = None,
    stats_lhm: dict | None = None,
    categorie: str = "",
    p95_klasse: str = "",
) -> None:
    """Maakt een figuur met twee subplots: decadewaarden en cumulatieve waarden.

    Bovenste plot : decade debiet [m³/s] voor meting, Ribasim en LHM 4.1.
    Onderste plot : cumulatieve afvoer [Mm³] voor alle drie.
    Rechts        : statistieken + toetsingsresultaten als tekstblok buiten de assen.
    """
    Path(output_folder).mkdir(parents=True, exist_ok=True)

    MULTIPLIER = 10 * 86400  # decade → seconden
    font = "Arial"

    datum = df_merged["Datum"]
    meting = df_merged["meting_som"]
    ribasim = df_merged["Ribasim_som"]
    lhm41 = df_merged["LHM41"]

    cum_meting = meting.fillna(0).cumsum() * MULTIPLIER / 1_000_000
    cum_ribasim = ribasim.fillna(0).cumsum() * MULTIPLIER / 1_000_000
    cum_lhm41 = lhm41.fillna(0).cumsum() * MULTIPLIER / 1_000_000

    heeft_stats = stats_rib is not None and stats_lhm is not None
    fig_breedte = 12 if heeft_stats else 8

    fig = plt.figure(figsize=(fig_breedte, 7))
    if heeft_stats:
        gs = fig.add_gridspec(2, 2, width_ratios=[3, 1.8], hspace=0.15, wspace=0.12)
        ax_dec = fig.add_subplot(gs[0, 0])
        ax_cum = fig.add_subplot(gs[1, 0], sharex=ax_dec)
        ax_leg = fig.add_subplot(gs[:, 1])
        ax_leg.axis("off")
    else:
        gs = fig.add_gridspec(2, 1, hspace=0.15)
        ax_dec = fig.add_subplot(gs[0])
        ax_cum = fig.add_subplot(gs[1], sharex=ax_dec)

    fig.suptitle(titel, fontname=font, fontsize=12, fontweight="bold", wrap=True)

    # ── Bovenste plot: decadewaarden ──────────────────────────────────────────
    ax_dec.plot(datum, meting, label="Meting", color="tab:orange", linewidth=2.0)
    ax_dec.plot(datum, ribasim, label="Ribasim", color="tab:blue", linewidth=2.0)
    ax_dec.plot(datum, lhm41, label="LHM 4.1", color="tab:green", linewidth=2.0, linestyle="--")
    ax_dec.set_ylabel("Debiet [m³/s]", fontdict={"fontsize": 12, "fontname": font, "fontweight": "bold"})
    ax_dec.legend(fontsize=10, loc="upper right")
    ax_dec.grid(True, linewidth=0.5)
    ax_dec.tick_params(axis="both", labelsize=11)
    for lbl in ax_dec.get_xticklabels() + ax_dec.get_yticklabels():
        lbl.set_fontweight("bold")

    # ── Onderste plot: cumulatieve waarden ────────────────────────────────────
    ax_cum.plot(datum, cum_meting, label="Meting cum.", color="tab:orange", linewidth=2.0)
    ax_cum.plot(datum, cum_ribasim, label="Ribasim cum.", color="tab:blue", linewidth=2.0)
    ax_cum.plot(datum, cum_lhm41, label="LHM 4.1 cum.", color="tab:green", linewidth=2.0, linestyle="--")
    ax_cum.set_ylabel("Cum. debiet [Mm³]", fontdict={"fontsize": 12, "fontname": font, "fontweight": "bold"})
    ax_cum.legend(fontsize=10, loc="upper left")
    ax_cum.grid(True, linewidth=0.5)
    ax_cum.tick_params(axis="x", rotation=45, labelsize=11)
    ax_cum.tick_params(axis="y", labelsize=11)
    for lbl in ax_cum.get_xticklabels() + ax_cum.get_yticklabels():
        lbl.set_fontweight("bold")

    # ── Rechts: statistieken + toetsing ──────────────────────────────────────
    # if heeft_stats:
    #     regels = _stats_regels(stats_rib, stats_lhm, categorie, p95_klasse)
    #     # Achtergrondkader
    #     ax_leg.add_patch(plt.matplotlib.patches.FancyBboxPatch(
    #         (0.02, 0.02), 0.96, 0.96,
    #         transform=ax_leg.transAxes,
    #         boxstyle="round,pad=0.02",
    #         facecolor="whitesmoke", edgecolor="lightgray", linewidth=0.8, zorder=0,
    #     ))
    #     LINE_HEIGHT = 0.072
    #     y = 0.96
    #     for regel in regels:
    #         tekst = regel["text"]
    #         kleur = regel["color"] or "black"
    #         gewicht = "bold" if regel.get("bold") else "normal"
    #         suffix = regel.get("suffix")
    #         suffix_kleur = regel.get("suffix_color", "black")
    #         ax_leg.text(
    #             0.06, y, tekst,
    #             transform=ax_leg.transAxes,
    #             fontsize=10.5, fontfamily=font, color=kleur,
    #             fontweight=gewicht, verticalalignment="top",
    #             clip_on=True,
    #         )
    #         if suffix:
    #             ax_leg.text(
    #                 0.96, y, suffix,
    #                 transform=ax_leg.transAxes,
    #                 fontsize=10.5, fontfamily=font, color=suffix_kleur,
    #                 fontweight="bold", verticalalignment="top",
    #                 ha="right", clip_on=True,
    #             )
    #         y -= LINE_HEIGHT

    if heeft_stats:
        regels = _stats_regels(stats_rib, stats_lhm, categorie, p95_klasse)
        # Achtergrondkader
        ax_leg.add_patch(
            plt.matplotlib.patches.FancyBboxPatch(
                (0.02, 0.02),
                0.96,
                0.96,
                transform=ax_leg.transAxes,
                boxstyle="round,pad=0.02",
                facecolor="whitesmoke",
                edgecolor="lightgray",
                linewidth=0.8,
                zorder=0,
            )
        )
        LINE_HEIGHT = 0.050
        VERDICT_STEP = 0.020
        GAP_AFTER = 0.030

        y = 0.96

        for regel in regels:
            tekst = regel["text"]
            kleur = regel["color"] or "black"
            suffix = regel.get("suffix")
            suffix_kleur = regel.get("suffix_color", "black")

            # Score / titel
            ax_leg.text(
                0.06,
                y,
                tekst,
                transform=ax_leg.transAxes,
                fontsize=10.5,
                fontfamily=font,
                color=kleur,
                fontweight="bold",
                verticalalignment="top",
                clip_on=True,
            )

            # Stap naar beneden voor volgende regel
            y -= LINE_HEIGHT

            if suffix:
                # 👉 DIRECT tekenen (niet eerst extra omlaag!)
                ax_leg.text(
                    0.08,
                    y,
                    suffix,
                    transform=ax_leg.transAxes,
                    fontsize=10.5,
                    fontfamily=font,
                    color=suffix_kleur,
                    fontweight="bold",
                    verticalalignment="top",
                    clip_on=True,
                )

                # Daarna pas spacing
                y -= VERDICT_STEP
                y -= GAP_AFTER

    fig.savefig(Path(output_folder) / f"{fig_naam}.png", dpi=300, bbox_inches="tight")
    plt.close(fig)


def _kleur_toetsing(ws, df: pd.DataFrame) -> None:
    """Kleurt de Beoordeling-kolom in het toetsing-worksheet."""
    if "Beoordeling" not in df.columns:
        return
    col_idx = df.columns.tolist().index("Beoordeling") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = BEOOR_KLEUREN.get(cell.value, BEOOR_KLEUREN["n.v.t."])


# ── Hoofd-functie ─────────────────────────────────────────────────────────────


def AnalyseLHM41Vergelijking(
    gpkg_koppellaag: str,
    layer_naam: str,
    model_folder: str,
    meas_folder: str,
    lhm41_folder: str,
    output_path: str | None = None,
) -> None:
    """Voert de LHM 4.1-vergelijkingsanalyse uit en schrijft een criteria-Excel weg.

    Parameters
    ----------
    gpkg_koppellaag
        Pad naar de geopackage met de koppellaag.
    layer_naam
        Naam van de laag in de geopackage (bijv. ``'koppeling_lhm4_1_reeksen'``).
    model_folder
        Map van het Ribasim-model; tijdreeksen worden gelezen uit
        ``{model_folder}/results/tijdreeksen/``.
    meas_folder
        Map met de meetbestanden (Metingen_aanvoer/afvoer_dag_totaal.csv).
    lhm41_folder
        Map met de LHM 4.1 decade CSV-bestanden.
    output_path
        Pad voor het output-Excel. Standaard:
        ``{model_folder}/results/Validatie_criteria_lhm4_1.xlsx``.
    """
    if output_path is None:
        output_path = Path(model_folder) / "results" / "Validatie_criteria_lhm4_1.xlsx"

    ts_root = Path(model_folder) / "results" / "tijdreeksen"

    # ── 1. Laad koppellaag ────────────────────────────────────────────────────
    print("Laden koppellaag...")
    gdf = gpd.read_file(gpkg_koppellaag, layer=layer_naam)
    ontbrekend = {"MeetreeksC", "Aan/Af", "Waterschap", "LHM_4_1_csv"} - set(gdf.columns)
    if ontbrekend:
        raise ValueError(f"Kolommen ontbreken in koppellaag: {ontbrekend}")
    print(f"  {len(gdf)} rijen geladen.")

    # ── 2. Laad metingen voor P95-classificatie ───────────────────────────────
    print("Laden metingen...")
    measurements = LoadMeasurements(meas_folder)

    # ── 3. Per locatie: decade-tijdreeks + statistieken Ribasim vs. meting ───
    print("Berekenen statistieken per locatie (Ribasim vs. meting)...")
    locatie_resultaten = []

    for _, rij in gdf.iterrows():
        meetreeks_c = rij["MeetreeksC"]
        aan_af = rij["Aan/Af"]
        waterschap = rij["Waterschap"]
        lhm41_csv = rij["LHM_4_1_csv"]
        categorie = _AANAF_NAAR_CATEGORIE.get(aan_af, "Afvoer")
        p95_klasse = _bereken_p95_klasse(measurements, meetreeks_c, aan_af)
        fig_name = _sanitize_filename(meetreeks_c)

        df_dec = _laad_decade_tijdreeks(ts_root, waterschap, fig_name)
        if df_dec is None or df_dec.dropna(subset=["meting", "Ribasim"]).empty:
            print(f"  Geen bruikbare decade-data voor {meetreeks_c}, overgeslagen.")
            continue

        stats_rib = _bereken_statistieken(df_dec, "Ribasim", "meting", categorie)

        locatie_resultaten.append(
            {
                "MeetreeksC": meetreeks_c,
                "Waterschap": waterschap,
                "Aan/Af": aan_af,
                "Categorie": categorie,
                "P95_klasse": p95_klasse,
                "LHM_4_1_csv": lhm41_csv,
                "df_dec": df_dec,
                "stats_rib": stats_rib,
            }
        )

    print(f"  {len(locatie_resultaten)} locaties verwerkt.")

    # ── 4. Groepeer op LHM_4_1_csv: sommeer + vergelijk met LHM 4.1 ─────────
    print("Berekenen statistieken per groep (LHM 4.1 vs. meting)...")
    groep_resultaten = []

    groepen: dict[str, list] = defaultdict(list)
    for loc in locatie_resultaten:
        csv_naam = str(loc["LHM_4_1_csv"]).strip()
        if csv_naam and csv_naam.lower() != "nan":
            groepen[csv_naam].append(loc)

    for csv_naam, locs in groepen.items():
        # Sommeer Ribasim en meting; datum-as van eerste tijdreeks is leidend
        datum_index = locs[0]["df_dec"]["Datum"].values
        rib_som = np.zeros(len(datum_index))
        met_som = np.zeros(len(datum_index))
        has_meting = np.zeros(len(datum_index), dtype=bool)

        for loc in locs:
            df_i = loc["df_dec"].set_index("Datum").reindex(datum_index)
            rib_som += df_i["Ribasim"].fillna(0).values
            met_som += df_i["meting"].fillna(0).values
            has_meting |= df_i["meting"].notna().values

        df_som = pd.DataFrame(
            {
                "Datum": datum_index,
                "Ribasim_som": rib_som,
                "meting_som": np.where(has_meting, met_som, np.nan),
            }
        )

        # LHM 4.1 inladen en mergen op datum (leidend = Ribasim tijdas)
        df_lhm = _laad_lhm41_csv(lhm41_folder, csv_naam)
        if df_lhm is None:
            continue
        df_merged = df_som.merge(df_lhm, on="Datum", how="left")

        categorie = locs[0]["Categorie"]
        p95_klasse = _bereken_p95_klasse_som(measurements, locs)
        if len(locs) > 1:
            individueel = [loc_["P95_klasse"] for loc_ in locs]
            if any(k != p95_klasse for k in individueel):
                namen = [loc_["MeetreeksC"] for loc_ in locs]
                print(
                    f"  P95-klasse gewijzigd na optelling voor {csv_naam}: "
                    f"individueel {individueel} → gecombineerd '{p95_klasse}' "
                    f"({', '.join(namen)})"
                )
        stats_lhm = _bereken_statistieken(df_merged, "LHM41", "meting_som", categorie)
        stats_rib_groep = _bereken_statistieken(df_merged, "Ribasim_som", "meting_som", categorie)

        # Figuur wegschrijven
        fig_naam = _sanitize_filename(Path(csv_naam).stem)
        titel = ", ".join(loc_["MeetreeksC"] for loc_ in locs)
        fig_folder = Path(model_folder) / "results" / "figures_lhm4_1_vergelijking"
        _plot_vergelijking(
            df_merged,
            titel,
            fig_folder,
            fig_naam,
            stats_rib=stats_rib_groep,
            stats_lhm=stats_lhm,
            categorie=categorie,
            p95_klasse=p95_klasse,
        )

        groep_resultaten.append(
            {
                "LHM_4_1_csv": csv_naam,
                "MeetreeksC": ", ".join(loc_["MeetreeksC"] for loc_ in locs),
                "Waterschap": locs[0]["Waterschap"],
                "Categorie": categorie,
                "P95_klasse": p95_klasse,
                "fig_naam": fig_naam,
                "locs": locs,
                "stats_rib": stats_rib_groep,
                "stats_lhm": stats_lhm,
            }
        )

    print(f"  {len(groep_resultaten)} groepen verwerkt.")

    # ── 5. Toetsing per (Categorie, P95_klasse) ───────────────────────────────
    print("Evalueren toetsingscriteria...")

    def _maak_toetsing_df(resultaten: list, stats_sleutel: str) -> pd.DataFrame:
        per_cat: dict[tuple, list] = defaultdict(list)
        for res in resultaten:
            per_cat[(res["Categorie"], res["P95_klasse"])].append(res[stats_sleutel])
        rijen = []
        for (cat, klasse), stats_lijst in sorted(per_cat.items()):
            rijen.extend(
                {"Categorie": cat, "P95_klasse": klasse, **tr} for tr in _toets_locaties(stats_lijst, cat, klasse)
            )
        return pd.DataFrame(rijen) if rijen else pd.DataFrame()

    df_toets_rib = _maak_toetsing_df(groep_resultaten, "stats_rib")
    df_toets_lhm = _maak_toetsing_df(groep_resultaten, "stats_lhm")

    # ── 6. Statistieken-overzicht ─────────────────────────────────────────────
    def _stats_rij(s: dict, prefix: str) -> dict:
        rij = {
            f"{prefix}_NSE": round(s.get("NSE", np.nan), 3),
            f"{prefix}_RelBias": round(s.get("RelBias", np.nan), 1),
            f"{prefix}_CumJaar_gem": round(s.get("CumJaar", np.nan), 1),
            f"{prefix}_CumJaar_totaal": round(s.get("CumJaar_totaal", np.nan), 1),
            f"{prefix}_CumZomer_gem": round(s.get("CumZomer", np.nan), 1),
            f"{prefix}_CumZomer_totaal": round(s.get("CumZomer_totaal", np.nan), 1),
        }
        # NSE per jaar als losse kolommen: Rib_NSE_2017, Rib_NSE_2018, ...
        for jaar, nse in sorted(s.get("NSE_per_jaar", {}).items()):
            rij[f"{prefix}_NSE_{jaar}"] = round(nse, 3)
        # CumJaar per jaar als losse kolommen: Rib_CumJaar_2017, ...
        for jaar, afwijk in sorted(s.get("CumJaar_per_jaar", {}).items()):
            rij[f"{prefix}_CumJaar_{jaar}"] = round(afwijk, 1)
        # CumZomer per jaar als losse kolommen: Rib_CumZomer_2017, ...
        for jaar, afwijk in sorted(s.get("CumZomer_per_jaar", {}).items()):
            rij[f"{prefix}_CumZomer_{jaar}"] = round(afwijk, 1)
        return rij

    df_locaties = pd.DataFrame(
        [
            {
                "LHM_4_1_csv": g["LHM_4_1_csv"],
                "MeetreeksC": g["MeetreeksC"],
                "Waterschap": g["Waterschap"],
                "Categorie": g["Categorie"],
                "P95_klasse": g["P95_klasse"],
                **_stats_rij(g["stats_rib"], "Rib"),
            }
            for g in groep_resultaten
        ]
    )

    df_groepen = pd.DataFrame(
        [
            {
                "LHM_4_1_csv": g["LHM_4_1_csv"],
                "MeetreeksC": g["MeetreeksC"],
                "Categorie": g["Categorie"],
                "P95_klasse": g["P95_klasse"],
                **_stats_rij(g["stats_lhm"], "LHM41"),
            }
            for g in groep_resultaten
        ]
    )

    # ── 7. Schrijf Excel ──────────────────────────────────────────────────────
    print(f"Schrijven Excel: {output_path}")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not df_locaties.empty:
            df_locaties.to_excel(writer, sheet_name="Statistieken_Ribasim", index=False)
        if not df_groepen.empty:
            df_groepen.to_excel(writer, sheet_name="Statistieken_LHM41", index=False)
        if not df_toets_rib.empty:
            df_toets_rib.to_excel(writer, sheet_name="Toetsing_Ribasim", index=False)
            _kleur_toetsing(writer.sheets["Toetsing_Ribasim"], df_toets_rib)
        if not df_toets_lhm.empty:
            df_toets_lhm.to_excel(writer, sheet_name="Toetsing_LHM41", index=False)
            _kleur_toetsing(writer.sheets["Toetsing_LHM41"], df_toets_lhm)

    # ── 8. GeoPackage wegschrijven voor HTML-viewer ───────────────────────────
    print("GeoPackage wegschrijven voor HTML-viewer...")
    geom_lookup = {row["MeetreeksC"]: row.geometry for _, row in gdf.iterrows() if row.geometry is not None}

    gpkg_records = []
    for res in groep_resultaten:
        # Eerste locatie in de groep met bekende geometrie
        geom = None
        for loc in res["locs"]:
            geom = geom_lookup.get(loc["MeetreeksC"])
            if geom is not None:
                break
        if geom is None:
            continue

        fig_naam = res["fig_naam"]
        s_rib = res["stats_rib"]
        s_lhm = res["stats_lhm"]

        def _r(val, dec=2):
            return round(float(val), dec) if val is not None and not np.isnan(float(val)) else None

        gpkg_records.append(
            {
                "koppelinfo": res["MeetreeksC"],
                "MeetreeksC": res["MeetreeksC"],
                "Waterschap": res["Waterschap"],
                "Categorie": res["Categorie"],
                "P95_klasse": res["P95_klasse"],
                "LHM_4_1_csv": res["LHM_4_1_csv"],
                "NSE_rib": _r(s_rib.get("NSE")),
                "CumJaar_rib": _r(s_rib.get("CumJaar"), 1),
                "CumZomer_rib": _r(s_rib.get("CumZomer"), 1),
                "RelBias_rib": _r(s_rib.get("RelBias"), 1),
                "NSE_lhm": _r(s_lhm.get("NSE")),
                "CumJaar_lhm": _r(s_lhm.get("CumJaar"), 1),
                "CumZomer_lhm": _r(s_lhm.get("CumZomer"), 1),
                "RelBias_lhm": _r(s_lhm.get("RelBias"), 1),
                "figure_path": f'<img src="../figures_lhm4_1_vergelijking/{fig_naam}.png" width=400 height=300>',
                "geometry": geom,
            }
        )

    if gpkg_records:
        gdf_out = gpd.GeoDataFrame(gpkg_records, geometry="geometry", crs=gdf.crs)
        gpkg_out = Path(model_folder) / "results" / "Validatie_resultaten_lhm41.gpkg"
        gdf_out.to_file(gpkg_out, driver="GPKG", layer="lhm41_vergelijking")
        print(f"  GeoPackage opgeslagen: {gpkg_out} ({len(gdf_out)} groepen)")
    else:
        print("  Geen groepen met geometrie gevonden, GeoPackage niet weggeschreven.")

    print("Klaar.")


# %%

if __name__ == "__main__":
    cloud = CloudStorage()

    waterboard = "Limburg"
    waterboard_model_versions = cloud.uploaded_models(authority=waterboard)
    latest_model_version = sorted(
        [i for i in waterboard_model_versions if i.model == waterboard],
        key=lambda x: getattr(x, "sorter", ""),
    )[-1]

    model_folder = cloud.joinpath(f"{waterboard}/modellen", latest_model_version.path_string)
    meas_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/meetreeksen_2026")
    lhm41_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/LHM4_1_reeksen")
    gpkg_koppellaag = cloud.joinpath(
        "Basisgegevens/resultaatvergelijking/Koppeltabel_LHM4_1/Koppeltabel_met_toegevoegde_data_v28_04_2026.gpkg"
    )
    layer_naam = "koppeling_lhm4_1_reeksen"

    # Synchroniseer model en meetreeksen vanuit de cloud
    # cloud.synchronize([model_folder, meas_folder, lhm41_folder, gpkg_koppellaag])
    cloud.synchronize([lhm41_folder, gpkg_koppellaag])

    output_path = Path(model_folder) / "results" / "Validatie_criteria_lhm4_1.xlsx"

    AnalyseLHM41Vergelijking(
        gpkg_koppellaag=gpkg_koppellaag,
        layer_naam=layer_naam,
        model_folder=model_folder,
        meas_folder=meas_folder,
        lhm41_folder=lhm41_folder,
        output_path=output_path,
    )

    # ── Upload resultaten naar de cloud ──
    UPLOAD_cloud = False

    if UPLOAD_cloud:
        results = Path(model_folder) / "results"
        if output_path.exists():
            cloud.upload_file(output_path)
        gpkg_lhm41 = results / "Validatie_resultaten_lhm41.gpkg"
        if gpkg_lhm41.exists():
            cloud.upload_file(gpkg_lhm41)
        figures_lhm41 = results / "figures_lhm4_1_vergelijking"
        if figures_lhm41.exists():
            cloud.upload_content(figures_lhm41)

# %%
