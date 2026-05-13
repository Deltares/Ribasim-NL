# %%
"""Validatieanalyse regionale modellen.

Loop over alle actieve waterschappen: vergelijk Ribasim-uitvoer met meetreeksen,
bereken statistieken, maak figuren en schrijf een interactieve HTML-viewer.
Optioneel: voer ook de LHM 4.1-vergelijkingsanalyse uit per waterschap.
"""

from pathlib import Path

from openpyxl.styles import PatternFill
from ribasim_nl.analyse_results import (
    AnalyseLHM41Vergelijking,
    BerekenModelEindbeoordeling,
    CompareOutputMeasurements,
    ExtraInfoToevoegenAllData,
)
from ribasim_nl.html_viewer import CreateHTMLViewer

from ribasim_nl import CloudStorage

# %%
# ── Gedeelde cloud-paden ──────────────────────────────────────────────────────

cloud = CloudStorage()
base_koppeltabel = cloud.joinpath("Basisgegevens/resultaatvergelijking/koppeltabel_2026")
meas_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/meetreeksen_2026")

# LHM 4.1-vergelijking gedeelde paden
lhm41_folder = cloud.joinpath("Basisgegevens/resultaatvergelijking/LHM4_1_reeksen")
gpkg_koppellaag = cloud.joinpath(
    "Basisgegevens/resultaatvergelijking/Koppeltabel_LHM4_1/Koppeltabel_met_toegevoegde_data_v28_04_2026.gpkg"
)
lhm41_layer_naam = "koppeling_lhm4_1_reeksen"

# %%
# ── Toetsingscriteria en drempelwaarden ──────────────────────────────────────
# Richtwaarden: Goed = enige klasse (geen Matig); anders Onvoldoende.
# NSE > 0.4, KGE > 0.5, |Bias| < 15%, P10/P90 binnen ±30%, P25/P75 binnen ±15%.
CRITERIA_GRENZEN = {
    "Bias": {"Goed": 15.0},
    "NSE": {"Goed": 0.4},
    "KGE": {"Goed": 0.5},
    "P10": {"Goed": 30.0},
    "P25": {"Goed": 15.0},
    "P75": {"Goed": 15.0},
    "P90": {"Goed": 30.0},
}
BEOOR_KLEUREN = {
    "Goed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Matig": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Onvoldoende": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "n.v.t.": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
# 6 kwantitatieve criteria; model voldoet als ≥ MIN_GROEPEN_VOLDOEN ervan "Goed" zijn.
# NSE én KGE tellen samen als één criterium (tenminste één moet voldoen).
CRITERIA_GROEPEN = {
    "Bias": {"kolommen": ["Beoor_Bias_dec"], "min_goed": 1},
    "NSE_KGE": {"kolommen": ["Beoor_NSE_dec", "Beoor_KGE_dec"], "min_goed": 1},
    "P10": {"kolommen": ["Beoor_P10_dec"], "min_goed": 1},
    "P25": {"kolommen": ["Beoor_P25_dec"], "min_goed": 1},
    "P75": {"kolommen": ["Beoor_P75_dec"], "min_goed": 1},
    "P90": {"kolommen": ["Beoor_P90_dec"], "min_goed": 1},
}
MIN_GROEPEN_VOLDOEN = 3  # ≥ 3 van de 6 criteria Goed
DREMPEL_HWS = 75.0  # HWS: 75% van locaties moet voldoen
DREMPEL_REGIONAAL = 50.0  # Regionaal: 50% van locaties moet voldoen
HWS_WATERSCHAP = "Rijkswaterstaat"

# %%
# LHM 4.1-toetsingscriteria
CRITERIA_LHM41 = {
    ("Afvoer", ">=5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumJaar", 25, 80, "Cum. jaarafvoer dif. < 25%"),
    ],
    ("Afvoer", "<5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumJaar", 33, 80, "Cum. jaarafvoer dif. < 33%"),
    ],
    ("Aanvoer", ">=5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumZomer", 25, 80, "Cum. zomer aanvoer dif. < 25%"),
    ],
    ("Aanvoer", "<5"): [
        ("NSE", 0.2, 80, "NSE > 0.2"),
        ("NSE", 0.5, 60, "NSE > 0.5"),
        ("CumZomer", 50, 80, "Cum. zomer aanvoer dif. < 50%"),
    ],
}
# Tekencorrectie per LHM 4.1 CSV-bestand (-1 als tekenconventie verschilt van Ribasim/meting)
LHM41_TEKEN_CORRECTIE: dict[str, float] = {
    "Aanvoer Gemaal Winsemius_LHM_reeks.csv": -1.0,
}

# %%
# ── RUN-flags: zet op False om een stap over te slaan ──────────────────────

RUN_COMPARE = True
RUN_LHM41 = True
RUN_EXTRA_INFO = True
RUN_EINDBEOORDELING = True
RUN_HTML_VIEWER = True
RUN_UPLOAD = False

# HTML-viewer opties
model_gpkg_in_html = False  # netwerk tonen (kan traag zijn bij grote modellen)
include_lhm41_html = RUN_LHM41  # LHM 4.1 vergelijkingslaag tonen (vereist RUN_LHM41)
include_fractie_html = True  # fractieplot tonen (vereist concentration.nc uitvoer)

# %%
# ── Per-waterschap configuratie ───────────────────────────────────────────────
# Zet "active" op True om een waterschap te verwerken.
# "koppeltabel" en "specifiek" zijn bestandsnamen relatief aan base_koppeltabel.

WATERSCHAP_CONFIG = {
    # "Limburg": {
    #     "active": False,
    #     "koppeltabel": "Transformed_koppeltabel_versie_Limburg_2026_4_0_Feedback_Verwerkt_HydroLogic.xlsx",
    #     "specifiek":   "Specifiek_bewerking_versieLimburg_2026_4_0.xlsx",
    # },
    "RijnenIJssel": {
        "active": True,
        "koppeltabel": "Transformed_koppeltabel_versie_RijnenIJssel_2026_5_1_Feedback_Verwerkt_HydroLogic.xlsx",
        "specifiek": "Specifiek_bewerking_versieRijnenIJssel_2026_5_1.xlsx",
    },
}

# %%
# ── Synchroniseer gedeelde paden éénmalig ─────────────────────────────────────

cloud.synchronize([meas_folder])

if RUN_LHM41:
    cloud.synchronize([lhm41_folder, gpkg_koppellaag])

# %%
# ── Hoofdloop ─────────────────────────────────────────────────────────────────

for waterschap, cfg in WATERSCHAP_CONFIG.items():
    if not cfg["active"]:
        continue

    print(f"\n{'=' * 60}")
    print(f"  Verwerken: {waterschap}")
    print(f"{'=' * 60}")

    # ── Paden per waterschap ──────────────────────────────────────────────────
    loc_koppeltabel = base_koppeltabel / cfg["koppeltabel"]
    loc_specifieke_bewerking = base_koppeltabel / cfg["specifiek"]

    waterboard_model_versions = cloud.uploaded_models(authority=waterschap)
    latest_model_version = sorted(
        [i for i in waterboard_model_versions if i.model == waterschap],
        key=lambda x: getattr(x, "sorter", ""),
    )[-1]
    model_folder = cloud.joinpath(f"{waterschap}/modellen", latest_model_version.path_string)

    cloud.synchronize([loc_koppeltabel, loc_specifieke_bewerking, model_folder])

    # Zoek het .toml bestand dynamisch op in de gesynchroniseerde modelfolder
    toml_path = next(Path(model_folder).glob("*.toml"), None)
    if toml_path is None:
        print(f"  FOUT: geen .toml bestand gevonden in {model_folder} — {waterschap} overgeslagen.")
        continue
    toml_naam = toml_path.name
    print(f"  Model toml: {toml_naam}")

    results_folder = Path(model_folder) / "results"

    # ── Vergelijk uitvoer met meetreeksen, bereken statistieken ──────────────
    if RUN_COMPARE:
        CompareOutputMeasurements(
            loc_koppeltabel=loc_koppeltabel,
            loc_specifics=loc_specifieke_bewerking,
            meas_folder=meas_folder,
            model_folder=model_folder,
            toml_file=toml_naam,
            filetype="flow",
            apply_for_water_authority=None,  #!TODO: let op Deze alleen aanzetten als je met een koppeltabel werkt voor het hele LHM
            # In dat geval kun je de prefix logica gebruiken, voorkeur heeft eigelijk als je de resultaten
            # gaat terugbrengen naar waterschapsmodellen een aparte losse koppeltabel mee te geven, omdat
            # De koppelingen in de gehele koppeltabel van het LHM soms niet de juiste prefix bevatten en dus is
            # het linkje in het waterschapsmodel misschien niet (goed) terug te vinden.
            # !TODO: kloppen de gebruikte prefixen ook nog wel voor de nieuwe LHM modellen ?
            criteria_grenzen=CRITERIA_GRENZEN,
            beoor_kleuren=BEOOR_KLEUREN,
            save_results_combined=True,
            output_is_feather=False,
            output_is_nc=True,
            resample_to_daily=True,
        )
    else:
        print(
            "LET OP: RUN_COMPARE is uitgeschakeld. De overige stappen (RUN_EXTRA_INFO, "
            "RUN_EINDBEOORDELING, RUN_LHM41, RUN_HTML_VIEWER) vereisen dat "
            "CompareOutputMeasurements al eerder succesvol is gedraaid."
        )

    # ── LHM 4.1-vergelijking ──────────────────────────────────────────────────
    if RUN_LHM41:
        AnalyseLHM41Vergelijking(
            gpkg_koppellaag=gpkg_koppellaag,
            layer_naam=lhm41_layer_naam,
            model_folder=model_folder,
            meas_folder=meas_folder,
            lhm41_folder=lhm41_folder,
            criteria_lhm41=CRITERIA_LHM41,
            LHM41_TEKEN_CORRECTIE=LHM41_TEKEN_CORRECTIE,
        )

    # ── Voeg P95/P05-klasse toe aan geopackages ───────────────────────────────
    if RUN_EXTRA_INFO:
        ExtraInfoToevoegenAllData(
            loc_koppeltabel=loc_koppeltabel,
            meas_folder=meas_folder,
            model_folder=model_folder,
            stat_cols=("abs_q95", "abs_q05"),
            threshold=5,
            new_col="P95_P05_alle_beschikbare_data",
            mode="any",
            above_operator=">=",
        )

    # ── Eindbeoordeling op basis van Validatie_criteria.xlsx ─────────────────
    if RUN_EINDBEOORDELING:
        BerekenModelEindbeoordeling(
            validatie_xlsx=results_folder / "Validatie_criteria.xlsx",
            output_xlsx=results_folder / "Eindbeoordeling_model.xlsx",
            criteria_groepen=CRITERIA_GROEPEN,
            min_groepen_voldoen=MIN_GROEPEN_VOLDOEN,
            drempel_hws=DREMPEL_HWS,
            drempel_regionaal=DREMPEL_REGIONAAL,
            hws_waterschap=HWS_WATERSCHAP,
            beoor_kleuren=BEOOR_KLEUREN,
        )

    # ── HTML-viewer genereren ─────────────────────────────────────────────────
    if RUN_HTML_VIEWER:
        CreateHTMLViewer(
            model_folder=model_folder,
            model_gpkg=model_gpkg_in_html,
            include_lhm41=include_lhm41_html,
            include_fractie=include_fractie_html,
        )

    # ── Upload resultaten naar de cloud ───────────────────────────────────────
    #!TODO: nog niet getest
    # if RUN_UPLOAD:
    #     for gpkg in [
    #         "Validatie_resultaten_all.gpkg",
    #         "Validatie_resultaten_dec_all.gpkg",
    #         "Validatie_resultaten.gpkg",
    #         "Validatie_resultaten_dec.gpkg",
    #         "Validatie_resultaten_lhm41.gpkg",
    #     ]:
    #         gpkg_path = results_folder / gpkg
    #         if gpkg_path.exists():
    #             cloud.upload_file(gpkg_path)

    #     for xlsx in [
    #         "Validatie_criteria.xlsx",
    #         "Eindbeoordeling_model.xlsx",
    #         "Validatie_criteria_lhm4_1.xlsx",
    #     ]:
    #         xlsx_path = results_folder / xlsx
    #         if xlsx_path.exists():
    #             cloud.upload_file(xlsx_path)

    #     for folder in ["figures", "figures_fracties", "figures_lhm4_1_vergelijking", "Validatieresultaten_HTML"]:
    #         folder_path = results_folder / folder
    #         if folder_path.exists():
    #             cloud.upload_content(folder_path)

# %%
